#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Generador del proyecto "Calculadora Eléctrica Pro"
# Ejecuta este script para crear toda la estructura de la aplicación web.

# Uso:
# python generar_proyecto.py
# El script creará la carpeta 'calculadora_electrica_pro' con todos los archivos necesarios.

import os
import shutil
import sys

PROJECT_NAME = "calculadora_electrica_pro"
PROJECT_DIR = os.path.join(os.getcwd(), PROJECT_NAME)

# Contenido de los archivos a generar
FILES = {}

# # ----------------------------------------------------------------------
# requirements.txt
# # ----------------------------------------------------------------------
FILES["requirements.txt"] = '''Flask==2.3.3
Flask-SQLAlchemy==3.1.1
Flask-Migrate==4.0.5
WTForms==3.0.1
openpyxl==3.1.2
reportlab==4.0.4
Pillow==10.0.0
python-dotenv==1.0.0
'''

# # ----------------------------------------------------------------------
# README.md
# # ----------------------------------------------------------------------
FILES["README.md"] = '''# Calculadora Eléctrica Pro

Aplicación web profesional para diseño y cálculo de paneles eléctricos, circuitos, cargas y sistema solar respaldo.

## Características
- Gestión de proyectos eléctricos (paneleras)
- CRUD completo de zonas, circuitos y artefactos
- Cálculos automáticos: corriente, caída de tensión, breaker recomendado, calibre de cable
- Tabla de cargas con factores de demanda
- Módulo de respaldo solar con comparativa EcoFlow vs tradicional
- Exportación a PDF y Excel (multi-hoja)
- Interfaz responsiva con paleta técnica azul eléctrico

## Instalación y ejecución

```bash
cd calculadora_electrica_pro
python -m venv venv
source venv/bin/activate        # Linux/Mac
# o
venv\\Scripts\\activate          # Windows

pip install -r requirements.txt
flask db init
flask db migrate -m "initial"
flask db upgrade
flask init-data
flask run
Abrir http://127.0.0.1:5000

Credenciales de prueba
No requiere autenticación. Crea proyectos libremente.

Uso
Crear un proyecto (panelera)

Agregar zonas (ej: Cocina, Iluminación)

En cada zona, agregar circuitos con artefactos y cantidades

Visualizar cálculos automáticos

Exportar reporte PDF/Excel

Ver recomendación solar basada en consumo total

Estructura generada
app/ - aplicación Flask

models.py - modelos SQLAlchemy

routes/ - rutas principales y API

services/ - lógica de cálculos eléctricos y solar

exports/ - generadores PDF y Excel

templates/ - plantillas Jinja2

static/ - CSS, JS, imágenes
'''

# ----------------------------------------------------------------------
# run.py (punto de entrada)
# ----------------------------------------------------------------------
FILES["run.py"] = '''from app import create_app

app = create_app()

if __name__ == '__main__':
    app.run(debug=True)
'''
FILES["app.py"] = '''from app import create_app

app = create_app()

if __name__ == '__main__':
    app.run(debug=True)
'''

# ----------------------------------------------------------------------
# app/__init__.py
# ----------------------------------------------------------------------
FILES["app/__init__.py"] = '''from flask import Flask
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
import os
import click

db = SQLAlchemy()
migrate = Migrate()

def create_app():
    app = Flask(__name__)
    app.config['SECRET_KEY'] = 'clave-segura-electrica-pro'
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + \
        os.path.join(app.instance_path, 'electrical.db')
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

    db.init_app(app)
    migrate.init_app(app, db)

    # Registrar blueprints
    from app.routes.main import main_bp
    from app.routes.api import api_bp
    app.register_blueprint(main_bp)
    app.register_blueprint(api_bp, url_prefix='/api')

    # Asegurar carpeta instance
    os.makedirs(app.instance_path, exist_ok=True)

    # Comando CLI para inicializar datos
    @app.cli.command('init-data')
    def init_data():
        from app.models import Appliance
        initial_appliances = [
            ('Bombillo LED 9W', 9, 'Iluminación'), ('Bombillo LED 15W', 15, 'Iluminación'),
            ('Tubo Fluorescente 40W', 40,
             'Iluminación'), ('Luminaria LED 50W', 50, 'Iluminación'),
            ('Reflector LED 100W', 100, 'Iluminación'), ('Nevera 150W', 150, 'Cocina'),
            ('Nevera/Compresor 350W', 350,
             'Cocina'), ('Microondas 1200W', 1200, 'Cocina'),
            ('Horno Eléctrico 2000W', 2000,
             'Cocina'), ('Licuadora 600W', 600, 'Cocina'),
            ('Tostadora 800W', 800, 'Cocina'), ('Cafetera 900W', 900, 'Cocina'),
            ('Lavavajillas 1800W', 1800,
             'Cocina'), ('Aire Acondicionado 1 Ton', 1200, 'Climatización'),
            ('Aire Acondicionado 1.5 Ton', 1800,
             'Climatización'), ('Aire Acondicionado 2 Ton', 2400, 'Climatización'),
            ('Ventilador de Techo 75W', 75,
             'Climatización'), ('TV 32" LED', 60, 'Entretenimiento'),
            ('TV 55" LED', 120, 'Entretenimiento'), ('Computadora Desktop',
             300, 'Entretenimiento'),
            ('Laptop 65W', 65, 'Entretenimiento'), ('Consola Videojuegos',
             200, 'Entretenimiento'),
            ('Impresora Laser', 400, 'Oficina/Industrial'), ('Servidor Rack',
             500, 'Oficina/Industrial'),
            ('Lavadora 500W', 500,
             'Lavandería'), ('Secadora Eléctrica', 3000, 'Lavandería'),
            ('Plancha 1000W', 1000, 'Lavandería'), ('Calentador de Agua', 1500, 'Otros'),
            ('Bomba de Agua 750W', 750, 'Otros'), ('Taladro 800W', 800, 'Otros'),
            ('Compresor de Aire', 1500, 'Otros'), ('Cargador VE Nivel 2', 7200, 'Otros')
        ]
        for name, watts, cat in initial_appliances:
            if not Appliance.query.filter_by(name=name).first():
                db.session.add(
                    Appliance(name=name, power_watts=watts, category=cat, is_custom=False))
        db.session.commit()
        click.echo("Datos iniciales de artefactos cargados.")

    return app
'''

# ----------------------------------------------------------------------
# app/models.py
# ----------------------------------------------------------------------
FILES["app/models.py"] = '''from app import db
from datetime import datetime

class Project(db.Model):
    __tablename__ = 'projects'
    id = db.Column(db.Integer, primary_key=True)
    client_name = db.Column(db.String(200), nullable=False)
    electrician = db.Column(db.String(200), nullable=False)
    job_number = db.Column(db.String(100))
    observations = db.Column(db.Text)
    # 120V,208V,220V,240V,277V,480V
    voltage_system = db.Column(db.String(10), default='220V')
    central_breaker_amps = db.Column(db.Float, default=100)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    zones = db.relationship('Zone', backref='project',
                            lazy=True, cascade='all, delete-orphan')

    def total_connected_load(self):
        total_w = 0
        for zone in self.zones:
            total_w += zone.total_watts()
        return total_w

    def recommended_central_breaker(self):
        total_i = self.total_connected_load() / (float(self.voltage_system.rstrip('V')) * 0.9)
        standard_breakers = [15,20,25,30,40,50,60,70,80,
            100,125,150,175,200,225,250,300,400,500,600]
        for b in standard_breakers:
            if b >= total_i:
                return b
        return 600

class Zone(db.Model):
    __tablename__ = 'zones'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    voltage = db.Column(db.String(10), default='220V')
    demand_factor = db.Column(db.Float, default=1.0)  # 0-1
    project_id = db.Column(db.Integer, db.ForeignKey(
        'projects.id'), nullable=False)

    circuits = db.relationship(
        'Circuit', backref='zone', lazy=True, cascade='all, delete-orphan')

    def total_watts(self):
        return sum(c.total_watts() for c in self.circuits)

    def total_watts_with_demand(self):
        return self.total_watts() * self.demand_factor

    def total_current(self):
        v = float(self.voltage.rstrip('V'))
        if v <= 0:
            return 0
        return self.total_watts_with_demand() / v

class Circuit(db.Model):
    __tablename__ = 'circuits'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    voltage_override = db.Column(db.String(10), nullable=True)
    power_factor = db.Column(db.Float, default=0.9)
    safety_factor = db.Column(db.Float, default=1.25)  # 125%
    distance = db.Column(db.Float, default=10.0)
    distance_unit = db.Column(db.String(10), default='m')  # 'm' o 'ft'
    wire_gauge_mm2 = db.Column(db.Float, nullable=True)
    wire_gauge_awg = db.Column(db.String(5), nullable=True)
    breaker_curve = db.Column(db.String(1), default='C')
    # horas de uso diario estimado
    daily_hours = db.Column(db.Float, default=4.0)
    zone_id = db.Column(db.Integer, db.ForeignKey('zones.id'), nullable=False)

    appliances = db.relationship(
        'CircuitAppliance', backref='circuit', lazy=True, cascade='all, delete-orphan')

    def total_watts(self):
        return sum(ca.appliance.power_watts * ca.quantity for ca in self.appliances)

    def get_voltage(self):
        if self.voltage_override:
            return float(self.voltage_override.rstrip('V'))
        return float(self.zone.voltage.rstrip('V'))

    def current_nominal(self):
        v = self.get_voltage()
        if v == 0:
            return 0
        return self.total_watts() / (v * self.power_factor)

    def current_corrected(self):
        return self.current_nominal() * self.safety_factor

    def voltage_drop_percent(self):
        # Para cobre, resistividad 0.0172 ohm·mm²/m
        rho = 0.0172
        v = self.get_voltage()
        if v == 0:
            return 0
        I = self.current_nominal()
        L = self.distance if self.distance_unit == 'm' else self.distance * 0.3048
        # Sección transversal del cable en mm²
        if self.wire_gauge_mm2:
            A = self.wire_gauge_mm2
        elif self.wire_gauge_awg:
            awg_to_mm2 = {'14': 2.08, '12': 3.31, '10': 5.26, '8': 8.37, '6': 13.3, '4': 21.2,
                '2': 33.6, '1': 42.4, '1/0': 53.5, '2/0': 67.4, '3/0': 85.0, '4/0': 107.2}
            A = awg_to_mm2.get(self.wire_gauge_awg, 2.08)
        else:
            A = 2.08  # 14 AWG por defecto
        if A <= 0:
            return 0
        deltaV = (2 * rho * L * I) / A
        return (deltaV / v) * 100

    def recommended_breaker(self):
        standard = [15,20,25,30,35,40,45,50,60,70,80,
            90,100,110,125,150,175,200,225,250,300]
        i_corr = self.current_corrected()
        for b in standard:
            if b >= i_corr:
                return b
        return 300

class Appliance(db.Model):
    __tablename__ = 'appliances'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    power_watts = db.Column(db.Float, nullable=False)
    category = db.Column(db.String(50))
    is_custom = db.Column(db.Boolean, default=False)

class CircuitAppliance(db.Model):
    __tablename__ = 'circuit_appliances'
    id = db.Column(db.Integer, primary_key=True)
    circuit_id = db.Column(db.Integer, db.ForeignKey(
        'circuits.id'), nullable=False)
    appliance_id = db.Column(db.Integer, db.ForeignKey(
        'appliances.id'), nullable=False)
    quantity = db.Column(db.Integer, default=1)
    appliance = db.relationship('Appliance')
'''

# ----------------------------------------------------------------------
# app/calculations.py
# ----------------------------------------------------------------------
FILES["app/calculations.py"] = '''# Funciones de cálculo auxiliares
def awg_to_mm2(awg):
conversion = {
'14': 2.08, '12': 3.31, '10': 5.26, '8': 8.37, '6': 13.3,
'4': 21.2, '2': 33.6, '1': 42.4, '1/0': 53.5, '2/0': 67.4,
'3/0': 85.0, '4/0': 107.2
}
return conversion.get(awg, 2.08)

def mm2_to_awg(mm2):
for awg, val in {'14':2.08, '12':3.31, '10':5.26, '8':8.37, '6':13.3, '4':21.2, '2':33.6, '1':42.4, '1/0':53.5, '2/0':67.4, '3/0':85.0, '4/0':107.2}.items():
if mm2 <= val:
return awg
return '4/0'
'''

# ----------------------------------------------------------------------
# app/solar.py
# ----------------------------------------------------------------------
FILES["app/solar.py"] = '''def recommend_solar(total_daily_kwh, autonomy_hours=8, losses=0.2, dod=0.8, peak_sun_hours=5):
    # Capacidad batería necesaria (kWh)
    battery_kwh = (total_daily_kwh * autonomy_hours) / 24 * (1 + losses) / dod
    battery_kwh = max(battery_kwh, total_daily_kwh * 0.5)  # mínimo media día

    # Paneles necesarios (kWp)
    panel_wp = 400  # potencia pico estándar por panel
    required_kwp = (total_daily_kwh / peak_sun_hours) * (1 + losses)
    num_panels = max(1, round(required_kwp * 1000 / panel_wp))

    inverter_power = total_daily_kwh / peak_sun_hours * 1000 * 1.25

    return {
        'battery_kwh': round(battery_kwh, 1),
        'num_panels': num_panels,
        'inverter_power_w': round(inverter_power, 0),
        'daily_consumption_kwh': round(total_daily_kwh, 2)
    }

def ecoflow_options(daily_kwh):
    options = []
    models = [
        ('RIVER 2 Pro', 0.768, 800),
        ('DELTA 2', 1.024, 1800),
        ('DELTA 2 Max', 2.048, 2400),
        ('DELTA Pro', 3.6, 3600),
        ('DELTA Pro Ultra', 6.144, 7200)
    ]
    for name, cap, inv in models:
        autonomy = cap / max(daily_kwh/24, 0.1)  # horas
        if autonomy > 0.5:
            options.append({
                'model': name,
                'capacity_kwh': cap,
                'inverter_w': inv,
                'autonomy_hours': round(autonomy, 1),
                'panels_needed': max(1, round(daily_kwh / 2.5))
            })
    return options[:4]
'''

# ----------------------------------------------------------------------
# app/exports/pdf_generator.py
# ----------------------------------------------------------------------
FILES["app/exports/pdf_generator.py"] = '''from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io

def generate_project_pdf(project, zones_data):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm,
                            leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    story = []

    # Título
    title_style = ParagraphStyle(
        'Title', parent=styles['Title'], alignment=TA_CENTER, fontSize=16, textColor=colors.darkblue)
    story.append(Paragraph("INFORME ELÉCTRICO PROFESIONAL", title_style))
    story.append(Spacer(1, 10*mm))

    # Datos del proyecto
    data = [
        ["Cliente:", project.client_name],
        ["Electricista:", project.electrician],
        ["N° Obra:", project.job_number or "-"],
        ["Voltaje sistema:", project.voltage_system],
        ["Breaker central:", f"{project.central_breaker_amps} A"],
        ["Observaciones:", project.observations or "-"]
    ]
    t = Table(data, colWidths=[50*mm, 100*mm])
    t.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey),
               ('BACKGROUND', (0,0), (0,-1), colors.lightgrey)]))
    story.append(t)
    story.append(Spacer(1, 10*mm))

    # Tabla de cargas por zona
    for zone in project.zones:
        story.append(Paragraph(
            f"Zona: {zone.name} - Voltaje {zone.voltage} - Factor Demanda {zone.demand_factor*100}%", styles['Heading4']))
        circ_data = [["Circuito", "Artefactos", "W Total",
            "I (A)", "Breaker (A)", "Cable", "Caída %"]]
        for circuit in zone.circuits:
            artifacts = ", ".join(
                [f"{ca.appliance.name} x{ca.quantity}" for ca in circuit.appliances])
            cable = circuit.wire_gauge_mm2 if circuit.wire_gauge_mm2 else circuit.wire_gauge_awg
            circ_data.append([
                circuit.name, artifacts[:50], f"{circuit.total_watts():.0f}",
                f"{circuit.current_nominal():.1f}", f"{circuit.recommended_breaker()}",
                cable or "2.08mm²", f"{circuit.voltage_drop_percent():.1f}%"
            ])
        t_circ = Table(circ_data, repeatRows=1)
        t_circ.setStyle(TableStyle(
            [('GRID', (0,0), (-1,-1), 0.5, colors.black), ('BACKGROUND', (0,0), (-1,0), colors.lightblue)]))
        story.append(t_circ)
        story.append(Spacer(1, 5*mm))

    story.append(Paragraph("Firmas:", styles['Normal']))
    story.append(Spacer(1, 20*mm))
    story.append(Paragraph("_________________________", styles['Normal']))
    story.append(Paragraph("Electricista", styles['Normal']))
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph("_________________________", styles['Normal']))
    story.append(Paragraph("Cliente", styles['Normal']))

    doc.build(story)
    buffer.seek(0)
    return buffer
'''

# ----------------------------------------------------------------------
# app/exports/excel_generator.py
# ----------------------------------------------------------------------
FILES["app/exports/excel_generator.py"] = '''import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io

def generate_excel_report(project):
    wb = openpyxl.Workbook()
    # Hoja 1: Proyecto
    ws1 = wb.active
    ws1.title = "Datos Proyecto"
    ws1['A1'] = "Calculadora Eléctrica Pro"
    ws1['A3'] = "Cliente:"; ws1['B3'] = project.client_name
    ws1['A4'] = "Electricista:"; ws1['B4'] = project.electrician
    ws1['A5'] = "N° Obra:"; ws1['B5'] = project.job_number or ""
    ws1['A6'] = "Voltaje sistema:"; ws1['B6'] = project.voltage_system
    ws1['A7'] = "Breaker central recomendado:"; ws1['B7'] = project.recommended_central_breaker()

    # Hoja 2: Tabla cargas
    ws2 = wb.create_sheet("Tabla Cargas")
    headers = ["Zona", "Circuito", "Artefactos", "W Total",
        "I (A)", "Breaker (A)", "Cable", "Caída %"]
    ws2.append(headers)
    for zone in project.zones:
        for circuit in zone.circuits:
            artifacts = ", ".join(
                [f"{ca.appliance.name} x{ca.quantity}" for ca in circuit.appliances])
            ws2.append([
                zone.name, circuit.name, artifacts, circuit.total_watts(),
                circuit.current_nominal(), circuit.recommended_breaker(),
                circuit.wire_gauge_mm2 or circuit.wire_gauge_awg or "2.08",
                circuit.voltage_drop_percent()
            ])

    # Hoja 3: Solar (placeholder, se puede mejorar)
    ws3 = wb.create_sheet("Sistema Solar")
    ws3.append(["Consumo diario estimado (kWh)", "Autonomía (h)",
               "Batería necesaria (kWh)", "Paneles (400W)", "Inversor (W)"])
    ws3.append(["---", "---", "---", "---", "---"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
'''

# ----------------------------------------------------------------------
# app/routes/main.py
# ----------------------------------------------------------------------
FILES["app/routes/main.py"] = '''from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from app import db
from app.models import Project, Zone, Circuit, CircuitAppliance, Appliance
from app.solar import recommend_solar, ecoflow_options
from app.exports.pdf_generator import generate_project_pdf
from app.exports.excel_generator import generate_excel_report
import io

main_bp = Blueprint('main', __name__)

@main_bp.route('/')
def index():
    projects = Project.query.all()
    return render_template('index.html', projects=projects)

@main_bp.route('/project/new', methods=['GET', 'POST'])
def new_project():
    if request.method == 'POST':
        p = Project(
            client_name=request.form['client_name'],
            electrician=request.form['electrician'],
            job_number=request.form.get('job_number'),
            observations=request.form.get('observations'),
            voltage_system=request.form['voltage_system'],
            central_breaker_amps=float(request.form['central_breaker_amps'])
        )
        db.session.add(p)
        db.session.commit()
        flash('Proyecto creado exitosamente', 'success')
        return redirect(url_for('main.project_detail', project_id=p.id))
    return render_template('project_form.html')

@main_bp.route('/project/<int:project_id>')
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)
    total_daily_kwh = 0
    for zone in project.zones:
        for circuit in zone.circuits:
            total_daily_kwh += circuit.total_watts() * circuit.daily_hours / 1000
    solar_rec = recommend_solar(total_daily_kwh)
    ecoflow_opts = ecoflow_options(total_daily_kwh)
    return render_template('project_detail.html', project=project, solar_rec=solar_rec, ecoflow_opts=ecoflow_opts)

@main_bp.route('/project/<int:project_id>/zone/new', methods=['GET', 'POST'])
def new_zone(project_id):
    project = Project.query.get_or_404(project_id)
    if request.method == 'POST':
        zone = Zone(
            name=request.form['name'],
            voltage=request.form['voltage'],
            demand_factor=float(request.form['demand_factor']),
            project_id=project.id
        )
        db.session.add(zone)
        db.session.commit()
        flash('Zona agregada', 'success')
        return redirect(url_for('main.project_detail', project_id=project.id))
    return render_template('zone_form.html', project=project)

@main_bp.route('/zone/<int:zone_id>/circuit/new', methods=['GET', 'POST'])
def new_circuit(zone_id):
    zone = Zone.query.get_or_404(zone_id)
    appliances = Appliance.query.all()
    if request.method == 'POST':
        circuit = Circuit(
            name=request.form['name'],
            voltage_override=request.form.get('voltage_override') or None,
            power_factor=float(request.form['power_factor']),
            safety_factor=float(request.form['safety_factor']),
            distance=float(request.form['distance']),
            distance_unit=request.form['distance_unit'],
            wire_gauge_mm2=float(request.form['wire_gauge_mm2']) if request.form.get('wire_gauge_mm2') else None,
            wire_gauge_awg=request.form.get('wire_gauge_awg') or None,
            breaker_curve=request.form['breaker_curve'],
            daily_hours=float(request.form['daily_hours']),
            zone_id=zone.id
        )
        db.session.add(circuit)
        db.session.commit()
        appliance_ids = request.form.getlist('appliance_id')
        quantities = request.form.getlist('quantity')
        for app_id, qty in zip(appliance_ids, quantities):
            if int(qty) > 0:
                ca = CircuitAppliance(circuit_id=circuit.id, appliance_id=int(app_id), quantity=int(qty))
                db.session.add(ca)
        db.session.commit()
        flash('Circuito creado', 'success')
        return redirect(url_for('main.project_detail', project_id=zone.project.id))
    return render_template('circuit_form.html', zone=zone, appliances=appliances)

@main_bp.route('/export/pdf/<int:project_id>')
def export_pdf(project_id):
    project = Project.query.get_or_404(project_id)
    pdf_buffer = generate_project_pdf(project, None)
    return send_file(pdf_buffer, as_attachment=True, download_name=f"proyecto_{project.id}.pdf", mimetype='application/pdf')

@main_bp.route('/export/excel/<int:project_id>')
def export_excel(project_id):
    project = Project.query.get_or_404(project_id)
    excel_buffer = generate_excel_report(project)
    return send_file(excel_buffer, as_attachment=True, download_name=f"proyecto_{project.id}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ====================== CRUD de Artefactos ======================
@main_bp.route('/appliances')
def list_appliances():
    appliances = Appliance.query.all()
    return render_template('appliances.html', appliances=appliances)

@main_bp.route('/appliance/new', methods=['GET', 'POST'])
def new_appliance():
    if request.method == 'POST':
        name = request.form['name']
        power_watts = float(request.form['power_watts'])
        category = request.form['category']
        appliance = Appliance(name=name, power_watts=power_watts, category=category, is_custom=True)
        db.session.add(appliance)
        db.session.commit()
        flash('Artefacto creado exitosamente', 'success')
        return redirect(url_for('main.list_appliances'))
    return render_template('appliance_form.html', title='Nuevo Artefacto', appliance=None)

@main_bp.route('/appliance/<int:id>/edit', methods=['GET', 'POST'])
def edit_appliance(id):
    appliance = Appliance.query.get_or_404(id)
    if request.method == 'POST':
        appliance.name = request.form['name']
        appliance.power_watts = float(request.form['power_watts'])
        appliance.category = request.form['category']
        db.session.commit()
        flash('Artefacto actualizado', 'success')
        return redirect(url_for('main.list_appliances'))
    return render_template('appliance_form.html', title='Editar Artefacto', appliance=appliance)

@main_bp.route('/appliance/<int:id>/delete', methods=['POST'])
def delete_appliance(id):
    appliance = Appliance.query.get_or_404(id)
    # Verificar si está siendo usado en algún circuito
    if appliance.circuit_appliances:
        flash('No se puede eliminar porque está siendo usado en circuitos', 'danger')
    else:
        db.session.delete(appliance)
        db.session.commit()
        flash('Artefacto eliminado', 'success')
    return redirect(url_for('main.list_appliances'))
'''
# ----------------------------------------------------------------------
# app/routes/api.py
# ----------------------------------------------------------------------
FILES["app/routes/api.py"] = '''from flask import Blueprint, jsonify, request
from app.models import Circuit

api_bp = Blueprint('api', __name__)


@api_bp.route('/circuit/<int:circuit_id>/calculate')
def calculate_circuit(circuit_id):
    circuit = Circuit.query.get_or_404(circuit_id)
    data = {
        'total_watts': circuit.total_watts(),
        'current_nominal': circuit.current_nominal(),
        'current_corrected': circuit.current_corrected(),
        'voltage_drop_percent': circuit.voltage_drop_percent(),
        'recommended_breaker': circuit.recommended_breaker()
    }
    return jsonify(data)
'''

# ----------------------------------------------------------------------
# Templates
# ----------------------------------------------------------------------
FILES["app/templates/base.html"] = '''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Calculadora Eléctrica Pro</title>
    <link href="{{ url_for('static', filename='css/bootstrap.min.css') }}" rel="stylesheet">
    <link rel="stylesheet" href="{{ url_for('static', filename='css/style.css') }}">
</head>
<body>
    <nav class="navbar navbar-dark bg-dark">
        <div class="container">
            <a class="navbar-brand" href="{{ url_for('main.index') }}">⚡ Calculadora Eléctrica Pro</a>
            <div>
                <a href="{{ url_for('main.list_appliances') }}" class="btn btn-outline-light btn-sm">Artefactos</a>
            </div>
        </div>
    </nav>
    <div class="container mt-4">
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }} alert-dismissible fade show" role="alert">
                        {{ message }}
                        <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        {% block content %}{% endblock %}
    </div>
    <script src="{{ url_for('static', filename='js/bootstrap.bundle.min.js') }}"></script>
    <script src="{{ url_for('static', filename='js/main.js') }}"></script>
</body>
</html>
'''

FILES["app/templates/index.html"] = '''{% extends 'base.html' %}
{% block content %}

<h2>Proyectos Eléctricos</h2> <a href="{{ url_for('main.new_project') }}" class="btn btn-primary mb-3">+ Nuevo Proyecto</a> <div class="row"> {% for project in projects %} <div class="col-md-4 mb-3"> <div class="card"> <div class="card-body"> <h5 class="card-title">{{ project.client_name }}</h5> <p class="card-text">Electricista: {{ project.electrician }}<br>Voltaje: {{ project.voltage_system }}</p> <a href="{{ url_for('main.project_detail', project_id=project.id) }}" class="btn btn-sm btn-info">Ver Detalle</a> </div> </div> </div> {% endfor %} </div> {% endblock %} '''
FILES["app/templates/project_detail.html"] = '''{% extends 'base.html' %}
{% block content %}

<div class="d-flex justify-content-between"> <h2>Panelera: {{ project.client_name }}</h2> <div> <a href="{{ url_for('main.export_pdf', project_id=project.id) }}" class="btn btn-danger">📄 PDF</a> <a href="{{ url_for('main.export_excel', project_id=project.id) }}" class="btn btn-success">📊 Excel</a> </div> </div> <p><strong>Electricista:</strong> {{ project.electrician }} | <strong>N° Obra:</strong> {{ project.job_number or '-' }} | <strong>Voltaje Sistema:</strong> {{ project.voltage_system }}</p> <a href="{{ url_for('main.new_zone', project_id=project.id) }}" class="btn btn-outline-primary mb-3">+ Agregar Zona</a>
{% for zone in project.zones %}

<div class="card mt-3"> <div class="card-header bg-secondary text-white"> <strong>{{ zone.name }}</strong> - Voltaje: {{ zone.voltage }} - Factor Demanda: {{ (zone.demand_factor*100)|int }}% </div> <div class="card-body"> <a href="{{ url_for('main.new_circuit', zone_id=zone.id) }}" class="btn btn-sm btn-success mb-2">+ Circuito</a> <table class="table table-sm table-bordered"> <thead> <tr><th>Circuito</th><th>Artefactos</th><th>W Total</th><th>I (A)</th><th>Breaker</th><th>Cable</th><th>Caída %</th><th>Estado</th></tr> </thead> <tbody> {% for circuit in zone.circuits %} <tr> <td>{{ circuit.name }}</td> <td>{% for ca in circuit.appliances %}{{ ca.appliance.name }} x{{ ca.quantity }}<br>{% endfor %}</td> <td>{{ circuit.total_watts()|round(0) }} W</td> <td>{{ circuit.current_nominal()|round(1) }} A</td> <td>{{ circuit.recommended_breaker() }} A</td> <td>{{ circuit.wire_gauge_mm2 or circuit.wire_gauge_awg or '2.08 mm²' }}</td> <td class="{% if circuit.voltage_drop_percent() > 3 %}text-danger{% else %}text-success{% endif %}">{{ circuit.voltage_drop_percent()|round(1) }}%</td> <td>{% if circuit.voltage_drop_percent() > 3 %}⚠️ Alerta{% else %}✅ OK{% endif %}</td> </tr> {% endfor %} </tbody> </table> <p><strong>Subtotal Zona:</strong> {{ zone.total_watts_with_demand()|round(0) }} W | {{ (zone.total_watts_with_demand() / (zone.voltage.rstrip('V')|float))|round(1) }} A</p> </div> </div> {% endfor %}<h4 class="mt-4">Total General</h4> <p><strong>Potencia total:</strong> {{ project.total_connected_load()|round(0) }} W</p> <p><strong>Breaker Central Recomendado:</strong> {{ project.recommended_central_breaker() }} A</p><h4>☀️ Sistema Solar Recomendado</h4> <p>Consumo diario estimado: {{ solar_rec.daily_consumption_kwh }} kWh/día</p> <p>Baterías necesarias: {{ solar_rec.battery_kwh }} kWh | Paneles 400W: {{ solar_rec.num_panels }} uds | Inversor: {{ solar_rec.inverter_power_w }} W</p> <h5>Opción EcoFlow</h5> <ul> {% for opt in ecoflow_opts %} <li>{{ opt.model }} - Capacidad {{ opt.capacity_kwh }} kWh, Autonomía {{ opt.autonomy_hours }} h, Paneles sugeridos: {{ opt.panels_needed }}</li> {% endfor %} </ul> {% endblock %} '''
FILES["app/templates/project_form.html"] = '''{% extends 'base.html' %}
{% block content %}

<h2>Nuevo Proyecto</h2> <form method="post"> <div class="mb-3"><label>Cliente</label><input type="text" name="client_name" class="form-control" required></div> <div class="mb-3"><label>Electricista</label><input type="text" name="electrician" class="form-control" required></div> <div class="mb-3"><label>N° Obra</label><input type="text" name="job_number" class="form-control"></div> <div class="mb-3"><label>Voltaje sistema</label><select name="voltage_system" class="form-select"><option>120V</option><option>208V</option><option>220V</option><option>240V</option><option>277V</option><option>480V</option></select></div> <div class="mb-3"><label>Breaker Central (A)</label><input type="number" step="5" name="central_breaker_amps" class="form-control" value="100"></div> <div class="mb-3"><label>Observaciones</label><textarea name="observations" class="form-control"></textarea></div> <button type="submit" class="btn btn-primary">Guardar</button> </form> {% endblock %} '''
FILES["app/templates/zone_form.html"] = '''{% extends 'base.html' %}
{% block content %}

<h2>Nueva Zona para {{ project.client_name }}</h2> <form method="post"> <div class="mb-3"><label>Nombre Zona</label><input type="text" name="name" class="form-control" required></div> <div class="mb-3"><label>Voltaje</label><select name="voltage" class="form-select"><option>120V</option><option>220V</option><option>240V</option></select></div> <div class="mb-3"><label>Factor Demanda (0.0-1.0)</label><input type="number" step="0.05" name="demand_factor" class="form-control" value="1.0"></div> <button type="submit" class="btn btn-primary">Crear Zona</button> </form> {% endblock %} '''
FILES["app/templates/circuit_form.html"] = '''{% extends 'base.html' %}
{% block content %}

<h2>Circuito en Zona: {{ zone.name }}</h2> <form method="post"> <div class="mb-3"><label>Nombre Circuito</label><input type="text" name="name" class="form-control" required></div> <div class="mb-3"><label>Voltaje (dejar vacío para heredar de zona)</label><input type="text" name="voltage_override" placeholder="Ej: 120V" class="form-control"></div> <div class="mb-3"><label>Factor Potencia</label><input type="number" step="0.01" name="power_factor" class="form-control" value="0.9"></div> <div class="mb-3"><label>Factor Seguridad</label><input type="number" step="0.05" name="safety_factor" class="form-control" value="1.25"></div> <div class="mb-3"><label>Distancia</label><input type="number" step="0.1" name="distance" class="form-control" value="10"></div> <div class="mb-3"><label>Unidad</label><select name="distance_unit" class="form-select"><option value="m">metros</option><option value="ft">pies</option></select></div> <div class="mb-3"><label>Calibre cable mm²</label><input type="number" step="0.01" name="wire_gauge_mm2" class="form-control"></div> <div class="mb-3"><label>Calibre AWG</label><input type="text" name="wire_gauge_awg" placeholder="14,12,10..." class="form-control"></div> <div class="mb-3"><label>Curva Breaker</label><select name="breaker_curve" class="form-select"><option>B</option><option>C</option><option>D</option></select></div> <div class="mb-3"><label>Horas uso diario (para cálculo solar)</label><input type="number" step="0.5" name="daily_hours" class="form-control" value="4"></div> <h5>Artefactos</h5> <table class="table"> <thead> <tr><th>Artefacto</th><th>Cantidad</th></tr> </thead> <tbody> {% for app in appliances %} <tr> <td>{{ app.name }} ({{ app.power_watts }}W)</td> <td><input type="number" name="quantity" value="0" min="0" class="form-control" style="width:80px"></td> <input type="hidden" name="appliance_id" value="{{ app.id }}"> </tr> {% endfor %} </tbody> </table> <button type="submit" class="btn btn-primary">Guardar Circuito</button> </form> {% endblock %} '''

FILES["app/templates/appliances.html"] = '''{% extends 'base.html' %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
    <h2>Catálogo de Artefactos</h2>
    <a href="{{ url_for('main.new_appliance') }}" class="btn btn-success">+ Nuevo Artefacto</a>
</div>
<table class="table table-striped table-bordered">
    <thead>
        <tr>
            <th>Nombre</th>
            <th>Potencia (W)</th>
            <th>Categoría</th>
            <th>Acciones</th>
        </tr>
    </thead>
    <tbody>
        {% for app in appliances %}
        <tr>
            <td>{{ app.name }}</td>
            <td>{{ app.power_watts }}</td>
            <td>{{ app.category }}</td>
            <td>
                <a href="{{ url_for('main.edit_appliance', id=app.id) }}" class="btn btn-sm btn-warning">✏️ Editar</a>
                <form method="post" action="{{ url_for('main.delete_appliance', id=app.id) }}" style="display:inline;" onsubmit="return confirm('¿Eliminar {{ app.name }}?')">
                    <button type="submit" class="btn btn-sm btn-danger">🗑️ Eliminar</button>
                </form>
            </td>
        </tr>
        {% endfor %}
    </tbody>
</table>
{% endblock %}
'''

FILES["app/templates/appliance_form.html"] = '''{% extends 'base.html' %}
{% block content %}
<h2>{{ title }}</h2>
<form method="post">
    <div class="mb-3">
        <label>Nombre del Artefacto</label>
        <input type="text" name="name" class="form-control" value="{{ appliance.name if appliance else '' }}" required>
    </div>
    <div class="mb-3">
        <label>Potencia (W)</label>
        <input type="number" step="0.01" name="power_watts" class="form-control" value="{{ appliance.power_watts if appliance else '' }}" required>
    </div>
    <div class="mb-3">
        <label>Categoría</label>
        <select name="category" class="form-select" required>
            <option value="Iluminación" {% if appliance and appliance.category == 'Iluminación' %}selected{% endif %}>Iluminación</option>
            <option value="Cocina" {% if appliance and appliance.category == 'Cocina' %}selected{% endif %}>Cocina</option>
            <option value="Climatización" {% if appliance and appliance.category == 'Climatización' %}selected{% endif %}>Climatización</option>
            <option value="Entretenimiento" {% if appliance and appliance.category == 'Entretenimiento' %}selected{% endif %}>Entretenimiento</option>
            <option value="Oficina/Industrial" {% if appliance and appliance.category == 'Oficina/Industrial' %}selected{% endif %}>Oficina/Industrial</option>
            <option value="Lavandería" {% if appliance and appliance.category == 'Lavandería' %}selected{% endif %}>Lavandería</option>
            <option value="Otros" {% if appliance and appliance.category == 'Otros' %}selected{% endif %}>Otros</option>
        </select>
    </div>
    <button type="submit" class="btn btn-primary">Guardar</button>
    <a href="{{ url_for('main.list_appliances') }}" class="btn btn-secondary">Cancelar</a>
</form>
{% endblock %}
'''

# ----------------------------------------------------------------------
# Static files
# ----------------------------------------------------------------------
FILES["app/static/css/style.css"] = '''body { background-color: #F8F9FA; }
.navbar { background-color: #0A1628 !important; }
.btn-primary { background-color: #F39C12; border-color: #F39C12; }
.btn-primary:hover { background-color: #e67e22; }
.card-header.bg-secondary { background-color: #2C3E50 !important; }
.text-danger { font-weight: bold; }
'''

FILES["app/static/js/main.js"] = "// Funcionalidad adicional para cálculos en tiempo real\nconsole.log('Calculadora Eléctrica Pro cargada');"

# ----------------------------------------------------------------------
# Crear estructura de directorios y archivos
# ----------------------------------------------------------------------
import subprocess
import sys

def run_command(cmd, cwd=None):
    '''Ejecuta un comando en el sistema y muestra su salida en tiempo real.'''
    print(f"\n> {cmd}")
    process = subprocess.Popen(cmd, shell=True, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='utf-8', errors='replace')
    for line in process.stdout:
        print(line, end='')
    process.wait()
    if process.returncode != 0:
        print(f"⚠️  El comando falló con código {process.returncode}")
        sys.exit(process.returncode)

def run_application():
    '''Ejecuta todos los pasos para instalar y lanzar la aplicación.'''
    os.chdir(PROJECT_DIR)
    
    # 1. Crear entorno virtual
    run_command("python -m venv venv")
    
    # Determinar rutas según el SO
    if sys.platform == "win32":
        python_exe = "venv\\Scripts\\python.exe"
        flask_exe = "venv\\Scripts\\flask.exe"
    else:
        python_exe = "venv/bin/python"
        flask_exe = "venv/bin/flask"
    
    # 2. Instalar dependencias
    run_command(f"{python_exe} -m pip install --upgrade pip")
    run_command(f"{python_exe} -m pip install -r requirements.txt")
    
    # 3. Inicializar migraciones (si no existe la carpeta migrations)
    if not os.path.exists("migrations"):
        run_command(f"{flask_exe} db init")
    run_command(f"{flask_exe} db migrate -m 'initial'")
    run_command(f"{flask_exe} db upgrade")
    
    # 4. Cargar datos iniciales de artefactos
    run_command(f"{flask_exe} init-data")
    
    # 5. Ejecutar la aplicación (se queda en primer plano)
    print("\n🚀 Levantando servidor Flask...")
    run_command(f"{flask_exe} run")

def download_bootstrap():
    import urllib.request
    # CSS
    css_url = "https://cdn.jsdelivr.net/npm/bootstrap@5.3.0-alpha1/dist/css/bootstrap.min.css"
    css_path = os.path.join(PROJECT_DIR, "app/static/css/bootstrap.min.css")
    if not os.path.exists(css_path):
        print(f"Descargando Bootstrap CSS...")
        urllib.request.urlretrieve(css_url, css_path)
        print("Bootstrap CSS descargado.")
    else:
        print("Bootstrap CSS ya existe localmente.")
    
    # JS (bundle con Popper)
    js_url = "https://cdn.jsdelivr.net/npm/bootstrap@5.3.0-alpha1/dist/js/bootstrap.bundle.min.js"
    js_path = os.path.join(PROJECT_DIR, "app/static/js/bootstrap.bundle.min.js")
    if not os.path.exists(js_path):
        print(f"Descargando Bootstrap JS...")
        urllib.request.urlretrieve(js_url, js_path)
        print("Bootstrap JS descargado.")
    else:
        print("Bootstrap JS ya existe localmente.")

# ----------------------------------------------------------------------
# Crear estructura de directorios y archivos (versión modificada)
# ----------------------------------------------------------------------
def create_project():
    if not os.path.exists(PROJECT_DIR):
        os.makedirs(PROJECT_DIR)
    else:
        print(f"El directorio {PROJECT_DIR} ya existe. Se conservará la carpeta venv y la base de datos.")

    for filepath, content in FILES.items():
        full_path = os.path.join(PROJECT_DIR, filepath)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        with open(full_path, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"Creado: {full_path}")

    download_bootstrap()  # <-- NUEVO

    print("\n✅ Proyecto generado exitosamente.")
    print("🚀 Iniciando instalación automática y ejecución...\n")
    run_application()

if __name__ == '__main__':
    create_project()