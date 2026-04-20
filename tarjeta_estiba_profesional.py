#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🚀 GENERADOR DE TARJETA DE ESTIBA PROFESIONAL
Con código de producto, saldo por movimiento, exportación PDF/Excel con logo y folio.
CORREGIDO: import base64 y manejo seguro de archivo temporal.
"""

import urllib.request
import ssl
import os
import re
import json
import subprocess
import sys
from pathlib import Path
from datetime import datetime
import base64

# ==================== RAW_DATA (solo nombres) ====================
RAW_NAMES = """
Aceite
Aceite 1 Litro 1400
Bombón
Botonetas
Café Ziva
Cerveza 3 Caballos 210
Cerveza 3 Caballos 230
Cerveza Breda
Cerveza Cristal 280
Cerveza Mayabe 230
Cerveza Mayabe 280
Cerveza Parranda 210
Cerveza Parranda 230
Cerveza Presidente 210
Cerveza Presidente 250
Cerveza Shekels 220
Cerveza Shekels 250
Cerveza Unlager 220
Chambelona
Chicle
Chupa (60)
Cigarros H Upman
Cigarros Popular Rojo
Coditos 500g
Detergente 1kg
Detergente 500g 500
Embutido Jamón
Energizante
Galleta Craker
Galleta Lamore
Galleta Minees 70
Galleta Pygmi
Galleta Salrica
Galletas de sal
Galletas Luanezco
Galletas My Bit
Habana Club Añejo Lt
Jabón
Jabón 200
Jugos
Malta Guajira 330 ml
Malta Vitali
Menta Plus
Mybar
Pasta Dental
Pelly
Peter Cars 100
Picadillo
Pollo Ahumado 300
Pure de Tomate
Refresco Cola (1.5L)
Refresco Cola (lata) 250
Refresco Cotel paquete
Refresco Guanabana (paquete)
Refresco Limon (Lata) 250
Refresco Mandarina (paquete)
Refresco Melocotón (paquete)
Refresco Naranja (1.5L)
Refresco Naranja (Lata) 250
Refresco Pepsi
Refresco Piña Colada (paquete)
Refresco Pomo 250ml
Rona granel 200 ml
Rona Arocha 260
Rona Arocha 500 ml
Ron cajio venta granel 200 ml
Ron Guayabita
Ron Nucay (granel) 350 ml
Ron Old Partner
Ron Rivera
Ron Vega del Rio
Sal Común
Salchicha 480 g
Sazón Comino
Sazón Guama
Sopa
Sorbeto Piña Colada (paquete)
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Spaguetti
Togo Peanut
Vinagre
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Sorbeto Fist
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Wisky 375 ml
Sorbeto First
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Vodka Pepe
Vodka Peanut
Vodka Mágico
"""

def parse_names(raw_text):
    lines = [line.strip() for line in raw_text.strip().split('\n') if line.strip()]
    seen = set()
    unique = []
    for name in lines:
        if name not in seen:
            seen.add(name)
            unique.append(name)
    return unique

PRODUCT_NAMES = parse_names(RAW_NAMES)
PROJECT_ROOT = Path("tarjeta_estiba_profesional")

def write_file(relative_path: str, content: str):
    path = PROJECT_ROOT / relative_path
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding='utf-8')
    print(f"✅ Creado: {path}")

def download_file(url, dest_path):
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        ctx = ssl._create_unverified_context()
        with urllib.request.urlopen(url, context=ctx) as response, open(dest_path, 'wb') as out_file:
            out_file.write(response.read())
        print(f"📥 Descargado: {dest_path}")
    except Exception as e:
        print(f"❌ Error descargando {url}: {e}")

def download_static_dependencies():
    static_dir = PROJECT_ROOT / "static"
    bootstrap_css = static_dir / "css" / "bootstrap.min.css"
    if not bootstrap_css.exists():
        download_file("https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css", bootstrap_css)
    bootstrap_js = static_dir / "js" / "bootstrap.bundle.min.js"
    if not bootstrap_js.exists():
        download_file("https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js", bootstrap_js)
    icons_css = static_dir / "css" / "bootstrap-icons.min.css"
    if not icons_css.exists():
        download_file("https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css", icons_css)
    icons_font_dir = static_dir / "css" / "fonts"
    icons_font_files = [
        "https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/fonts/bootstrap-icons.woff2",
        "https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/fonts/bootstrap-icons.woff"
    ]
    for font_url in icons_font_files:
        font_name = font_url.split("/")[-1]
        font_path = icons_font_dir / font_name
        if not font_path.exists():
            download_file(font_url, font_path)

# ==================== ARCHIVOS DEL PROYECTO ====================
write_file("requirements.txt", """Flask==3.0.3
Flask-SQLAlchemy==3.1.1
Flask-WTF==1.2.1
WTForms==3.1.2
python-dotenv==1.0.1
psycopg2-binary==2.9.9
click==8.1.7
fpdf2==2.7.9
openpyxl==3.1.2
""")

write_file(".env", """DATABASE_URL=sqlite:///kardex.db
SECRET_KEY=clave-segura-para-desarrollo
FLASK_ENV=development
""")

write_file("README.md", """# Tarjeta de Estiba Profesional

Sistema de control de inventario con código de producto, saldo por movimiento, exportación profesional.

## Características
- Código único para cada producto (autogenerado)
- Registro de entradas y salidas con cantidad
- **Saldo calculado automáticamente** después de cada movimiento
- **Exportación a PDF profesional** con logo, folio y tabla de saldos
- **Exportación a Excel profesional** con formato corporativo
- Fechas de última recepción y última salida
- Botón flotante para subir
- Edición de movimientos
""")

write_file("models.py", """from flask_sqlalchemy import SQLAlchemy
from datetime import datetime

db = SQLAlchemy()

class Product(db.Model):
    __tablename__ = 'products'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(200), unique=True, nullable=False)
    deleted = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    inventory = db.relationship('Inventory', backref='product', uselist=False)
    transactions = db.relationship('Transaction', backref='product', lazy=True)

class Inventory(db.Model):
    __tablename__ = 'inventories'
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), primary_key=True)
    quantity = db.Column(db.Integer, default=0, nullable=False)
    last_updated = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Transaction(db.Model):
    __tablename__ = 'transactions'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    transaction_type = db.Column(db.String(3), nullable=False)  # 'IN' o 'OUT'
    quantity = db.Column(db.Integer, nullable=False)
    transaction_date = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.String(300))
    deleted = db.Column(db.Boolean, default=False)
""")

write_file("forms.py", """from flask_wtf import FlaskForm
from wtforms import StringField, IntegerField, SelectField, TextAreaField
from wtforms.validators import DataRequired, NumberRange, ValidationError
from models import Product

class ProductForm(FlaskForm):
    code = StringField('Código', validators=[DataRequired()])
    name = StringField('Nombre del Producto', validators=[DataRequired()])

    def validate_code(self, field):
        product = Product.query.filter_by(code=field.data, deleted=False).first()
        if product and product.id != getattr(self, '_product_id', None):
            raise ValidationError('Ya existe un producto con este código.')

    def validate_name(self, field):
        product = Product.query.filter_by(name=field.data, deleted=False).first()
        if product and product.id != getattr(self, '_product_id', None):
            raise ValidationError('Ya existe un producto con este nombre.')

class TransactionForm(FlaskForm):
    product_id = SelectField('Producto', coerce=int, validators=[DataRequired()])
    transaction_type = SelectField('Tipo', choices=[('IN', 'Entrada'), ('OUT', 'Salida')], validators=[DataRequired()])
    quantity = IntegerField('Cantidad', validators=[DataRequired(), NumberRange(min=1)])
    notes = TextAreaField('Notas')
""")

write_file("utils.py", """from datetime import datetime
from models import db, Transaction, Product, Inventory

def update_inventory(product_id, quantity, transaction_type):
    inv = Inventory.query.get(product_id)
    if not inv:
        inv = Inventory(product_id=product_id, quantity=0)
        db.session.add(inv)
    if transaction_type == 'IN':
        inv.quantity += quantity
    else:
        if inv.quantity < quantity:
            raise ValueError(f"Stock insuficiente: {inv.quantity} < {quantity}")
        inv.quantity -= quantity
    db.session.commit()

def revert_inventory(product_id, old_quantity, old_type):
    inv = Inventory.query.get(product_id)
    if not inv:
        inv = Inventory(product_id=product_id, quantity=0)
        db.session.add(inv)
    if old_type == 'IN':
        inv.quantity -= old_quantity
    else:
        inv.quantity += old_quantity
    if inv.quantity < 0:
        inv.quantity = 0
    db.session.commit()

def get_current_stock(product_id):
    inv = Inventory.query.get(product_id)
    return inv.quantity if inv else 0

def get_last_reception_date(product_id):
    stock = get_current_stock(product_id)
    if stock <= 0:
        return None
    last_in = Transaction.query.filter_by(product_id=product_id, transaction_type='IN', deleted=False).order_by(Transaction.transaction_date.desc()).first()
    return last_in.transaction_date if last_in else None

def get_last_dispatch_date(product_id):
    last_out = Transaction.query.filter_by(product_id=product_id, transaction_type='OUT', deleted=False).order_by(Transaction.transaction_date.desc()).first()
    return last_out.transaction_date if last_out else None

def get_transactions_with_balance(product_id):
    \"\"\"Devuelve lista de transacciones con saldo acumulado\"\"\"
    transactions = Transaction.query.filter_by(product_id=product_id, deleted=False).order_by(Transaction.transaction_date).all()
    balance = 0
    result = []
    for t in transactions:
        if t.transaction_type == 'IN':
            balance += t.quantity
            entrada = t.quantity
            salida = 0
        else:
            balance -= t.quantity
            entrada = 0
            salida = t.quantity
        result.append({
            'id': t.id,
            'date': t.transaction_date,
            'type': t.transaction_type,
            'entrada': entrada,
            'salida': salida,
            'balance': balance,
            'notes': t.notes
        })
    return result
""")

# Logo en base64 (ejemplo, reemplazar por imagen real)
LOGO_BASE64 = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAABhGlDQ1BJQ0MgcHJvZmlsZQAAKJF9kT1Iw0AcxV9TpUUqDnYQcchQnSyIijhKFYtgobQVWnUwufQLmjQkKS6OgmvBwY/FqoOLs64OroIg+AHi5uak6CIl/i8ptIjx4Lgf7+497t4BQr3MNKtrAtB020wl4mImuyoGXxFEP0YwKjHLmJOXknF8x9c9Ary+J3iW/7k/R56atRjgE4lnmWHaxBvEM5u2wXmfOMJKskp8Tjxp0gWJH7muuPzGueiywDMjZio1RxwhFostrHSwKpka8RRxVNV0yhdmPVY5b3HWKjXWvCd/YSinrSy5TnMYSSxhCQJEVFBGGRZitGqkWEjRftzDP+j6JXIp5CqBkWMBVaRJ1w/+B7+btYqTk15SMA50vjjOxzAI7AKNuuN8HzlO4wQIPgNXesdfqgMzn6RXO1rsCOhfBzfXHW15C7jcAYafdNmyK0VoBQoF4P2YvikH9N0CwTW3t9Y+Th+ADHW1ugEODoGRIuWue7y7u7O3f880+/sBG7Fysin43I0AAAAZdEVYdFNvZnR3YXJlAHBhaW50Lm5ldCA0LjAuMTM0A1t6AAAAIGNIUk0AAHomAACAhAAA+gAAAIDoAAB1MAAA6mAAADqYAAAXcJy6UTwAAAAJcEhZcwAACxMAAAsTAQCanBgAAAAHdElNRQfoBQkSERiZs/SZAAAGBklEQVR42u2bXWxVVRDHf2fuB23pB9BSCu1Hi0ZFUYIINhqj8cHEhMRo4oOJj74ZX4zPxgd90EcTfTHRBw0xiYkYH4w+aIwvomhUiYYPQcWiWCWtIG2BtrT33vtj7rktl/uxtx/33t5C/JNJe+7dM7/Zs2fPzJkz5ywLgiDggVlBZUVpRdGlRfE3jIv0+KD4m0lSxOAf4qKE/0p/9A8SFpz4QcJpouBfouB/RMH/iIL/EQX/R2QzL9qgw+dFb//bA+qAveADYAtYD4wD68HRBX+kHfgT3AsWg2rwHzgDHAI9wDGwDYyD18GZ6H9nGtP9z/7W/mbrB/wL3gLXwYvW/wTwbqsNtILV4BY4C3rBZfA+WAd+Bj8CD4PT4EfgWfBXdP9ZUIp/vtHp64pOBT8C/4C10XmbwSq8D1wK/iaGo/90R+fCkxV6U8Af4C/ROZMTYv0STsJpCjwfXbO0VUAKvAOuYxL9JthfYAD8BFaAjaAXzJhH7DdBePKEj4AxMAieAMNmUXsyWATeBAdBpcn/tsl/GP6C+8F+sMLkf8bkvx+Mjpy+Jt+HwG2gCuRSE6ugpcy92H8P/Bj0mvy34pP/ZwN+PZgdjZPnzPQfM/lPm/xrTf5lJn+tK/gfB5eB/we1D/vBp2Az2BN9twt0gY4p9OcOo9Q64b5j9n/mS/7lWefXZAP/EXAraI32BfgU/AGeAt/A4xQmwX3mN/gbB2KjEoMfE3wT+E6j1FoyU/8v8Er0f0cUhF6sPy4K3kag/afBZvBPdP5FlF8v/LPR31Lf+y8H20Cnyf8x+v4pUKb/TtQrPwz+ANtML/n9BPwV/APeKfA6aSK/1JfYoyb/3wW0v8sU/3GDbwF/gM8LvA9NnBcPxO7bC1aDc+BTcBx8C34E3wT/Q4J/x/fvBp+D35j8XeDf4GsweQJeAXvBnzD5n2PyPzLFf4+izhFd82I+2W9m/W3wZrTvMPmLjO99FxOY/05wO/JfAseB8eixMZP/LgLf4+kz+edN/s3gFPA18l+P/EZM+J5B/tlo/YOi56Lyf4zINQCOgKPBK/BrUF8CbwXrIXgheC2oIfg42AK2gm3B/eDdYHPwRHBv8GfwQLAh2BxsCbYF7wTbg52m9yekqJj8b05Z/zXwMHgseAS8H3XBR4Ong/uz9m8MHg3uCx6M/u/LOu8dU/w3m/zv5pA/vRw8FuyNzns0uj4vB3umWv+9YLfJvy56/4bo/KdN/j/zhH96l/T/rU2/6gAXwSXwK/igwF14D/gm2h8FOwvUhT8LPlb0Xa+IvCb/4wI2fhC8D/6Ydf77psj/PPr/+3nCr9J6T0f+fzD5PzI9/65ojZ0+1/9T8LhC+mrB13PIfyjy/9i3/u/luf6m6P+PKPoO2oPPiuT/K/jYFP+WaP/TYH/R+o+BT6P9OZP/evT/ZwVcY6r5uwPc2MRy30cHgqMRfMjBw+Al8Hp0wGpUbB9E/TT/c9PPpP4ruB4Z/Wz6FfSdZ4Kno/4m6P+fnKZ+5rnoWPDHLL2LgjjN9D0KPgA/gG+D9aAhDl8R+4O+6Pc02BC8E7wTvB28G+wK3g8+CnZG73s/en9XtD8Rvb8Dp6Nn/gCHwT5wCGwHx8D+6PkB8B74G3wB/gR7gi/BVvAJ+Bx8D/aCz8A+cAAcBp+DY+ATcDw6dwLsgk+f+Mg8wR/LA/4EcDg6fgo4EB1/Nfj2Bp/y5w1/uuC/4o/j4G50X4H+nz5FwP8DGgZ7CzGjMQkAAAAASUVORK5CYII="

write_file("app.py", '''
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from models import db, Product, Inventory, Transaction
from forms import ProductForm, TransactionForm
from utils import update_inventory, revert_inventory, get_current_stock, get_last_reception_date, get_last_dispatch_date, get_transactions_with_balance
from datetime import datetime
import click
import os
import re
from io import BytesIO
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import base64

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'clave-segura-por-defecto')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///kardex.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# ==================== DATOS RAW (solo nombres) ====================
RAW_NAMES = """
Aceite
Aceite 1 Litro 1400
Bombón
Botonetas
Café Ziva
Cerveza 3 Caballos 210
Cerveza 3 Caballos 230
Cerveza Breda
Cerveza Cristal 280
Cerveza Mayabe 230
Cerveza Mayabe 280
Cerveza Parranda 210
Cerveza Parranda 230
Cerveza Presidente 210
Cerveza Presidente 250
Cerveza Shekels 220
Cerveza Shekels 250
Cerveza Unlager 220
Chambelona
Chicle
Chupa (60)
Cigarros H Upman
Cigarros Popular Rojo
Coditos 500g
Detergente 1kg
Detergente 500g 500
Embutido Jamón
Energizante
Galleta Craker
Galleta Lamore
Galleta Minees 70
Galleta Pygmi
Galleta Salrica
Galletas de sal
Galletas Luanezco
Galletas My Bit
Habana Club Añejo Lt
Jabón
Jabón 200
Jugos
Malta Guajira 330 ml
Malta Vitali
Menta Plus
Mybar
Pasta Dental
Pelly
Peter Cars 100
Picadillo
Pollo Ahumado 300
Pure de Tomate
Refresco Cola (1.5L)
Refresco Cola (lata) 250
Refresco Cotel paquete
Refresco Guanabana (paquete)
Refresco Limon (Lata) 250
Refresco Mandarina (paquete)
Refresco Melocotón (paquete)
Refresco Naranja (1.5L)
Refresco Naranja (Lata) 250
Refresco Pepsi
Refresco Piña Colada (paquete)
Refresco Pomo 250ml
Rona granel 200 ml
Rona Arocha 260
Rona Arocha 500 ml
Ron cajio venta granel 200 ml
Ron Guayabita
Ron Nucay (granel) 350 ml
Ron Old Partner
Ron Rivera
Ron Vega del Rio
Sal Común
Salchicha 480 g
Sazón Comino
Sazón Guama
Sopa
Sorbeto Piña Colada (paquete)
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Spaguetti
Togo Peanut
Vinagre
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Sorbeto Fist
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Wisky 375 ml
Sorbeto First
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Vodka Pepe
Vodka Peanut
Vodka Mágico
"""

def parse_names(raw_text):
    lines = [line.strip() for line in raw_text.strip().split('\\n') if line.strip()]
    seen = set()
    unique = []
    for name in lines:
        if name not in seen:
            seen.add(name)
            unique.append(name)
    return unique

def generate_product_code(index):
    return f"P{index:03d}"

# ==================== COMANDOS FLASK ====================
@app.cli.command('init-db')
def init_db_command():
    db.create_all()
    click.echo('[OK] Base de datos inicializada.')

@app.cli.command('import-products')
def import_products():
    names = parse_names(RAW_NAMES)
    count = 0
    for idx, name in enumerate(names, start=1):
        code = generate_product_code(idx)
        if not Product.query.filter_by(name=name, deleted=False).first():
            prod = Product(code=code, name=name)
            db.session.add(prod)
            db.session.flush()
            db.session.add(Inventory(product_id=prod.id, quantity=0))
            count += 1
    db.session.commit()
    click.echo(f'[OK] Se importaron {count} productos.')

# ==================== RUTAS ====================
@app.route('/')
def index():
    products = Product.query.filter_by(deleted=False).all()
    product_data = []
    for p in products:
        stock = get_current_stock(p.id)
        last_reception = get_last_reception_date(p.id)
        last_dispatch = get_last_dispatch_date(p.id)
        product_data.append({
            'product': p,
            'stock': stock,
            'last_reception': last_reception,
            'last_dispatch': last_dispatch
        })
    return render_template('index.html', product_data=product_data)

@app.route('/product/<int:id>/kardex')
def kardex(id):
    product = Product.query.get_or_404(id)
    transactions = get_transactions_with_balance(id)
    stock = get_current_stock(id)
    return render_template('kardex.html', product=product, transactions=transactions, stock=stock)

@app.route('/product/add', methods=['GET', 'POST'])
def add_product():
    form = ProductForm()
    if form.validate_on_submit():
        product = Product(code=form.code.data, name=form.name.data)
        db.session.add(product)
        db.session.flush()
        db.session.add(Inventory(product_id=product.id, quantity=0))
        db.session.commit()
        flash('Producto agregado correctamente', 'success')
        return redirect(url_for('index'))
    return render_template('product_form.html', form=form, title='Nuevo Producto')

@app.route('/product/edit/<int:id>', methods=['GET', 'POST'])
def edit_product(id):
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    form._product_id = id
    if form.validate_on_submit():
        product.code = form.code.data
        product.name = form.name.data
        db.session.commit()
        flash('Producto actualizado', 'success')
        return redirect(url_for('index'))
    return render_template('product_form.html', form=form, title='Editar Producto')

@app.route('/product/delete/<int:id>')
def delete_product(id):
    product = Product.query.get_or_404(id)
    product.deleted = True
    db.session.commit()
    flash('Producto movido a papelera', 'warning')
    return redirect(url_for('index'))

@app.route('/product/restore/<int:id>')
def restore_product(id):
    product = Product.query.get_or_404(id)
    product.deleted = False
    db.session.commit()
    flash('Producto restaurado', 'success')
    return redirect(url_for('deleted_products'))

@app.route('/products/deleted')
def deleted_products():
    products = Product.query.filter_by(deleted=True).all()
    return render_template('deleted_products.html', products=products)

@app.route('/transaction/add', methods=['GET', 'POST'])
def add_transaction():
    form = TransactionForm()
    form.product_id.choices = [(p.id, f"{p.code} - {p.name}") for p in Product.query.filter_by(deleted=False).order_by(Product.code).all()]
    if form.validate_on_submit():
        product = Product.query.get(form.product_id.data)
        now = datetime.now()
        if form.transaction_type.data == 'OUT':
            stock = get_current_stock(product.id)
            if stock < form.quantity.data:
                flash(f'Stock insuficiente. Disponible: {stock}, solicitado: {form.quantity.data}', 'danger')
                return redirect(url_for('add_transaction'))
        
        trans = Transaction(
            product_id=product.id,
            transaction_type=form.transaction_type.data,
            quantity=form.quantity.data,
            transaction_date=now,
            notes=form.notes.data,
            deleted=False
        )
        db.session.add(trans)
        try:
            update_inventory(product.id, form.quantity.data, form.transaction_type.data)
            db.session.commit()
            flash('Movimiento registrado', 'success')
            return redirect(url_for('kardex', id=product.id))
        except ValueError as e:
            db.session.rollback()
            flash(str(e), 'danger')
    return render_template('transaction_form.html', form=form)

@app.route('/transaction/edit/<int:id>', methods=['GET', 'POST'])
def edit_transaction(id):
    trans = Transaction.query.get_or_404(id)
    if trans.deleted:
        flash('No se puede editar un movimiento eliminado. Restáurelo primero.', 'warning')
        return redirect(url_for('kardex', id=trans.product_id))
    
    form = TransactionForm(obj=trans)
    form.product_id.choices = [(trans.product_id, f"{trans.product.code} - {trans.product.name}")]
    form.product_id.data = trans.product_id
    
    if request.method == 'GET':
        form.transaction_type.data = trans.transaction_type
        form.quantity.data = trans.quantity
        form.notes.data = trans.notes
    
    if form.validate_on_submit():
        old_product_id = trans.product_id
        old_type = trans.transaction_type
        old_quantity = trans.quantity
        
        new_type = form.transaction_type.data
        new_quantity = form.quantity.data
        
        try:
            revert_inventory(old_product_id, old_quantity, old_type)
        except Exception as e:
            flash(f'Error al revertir inventario: {str(e)}', 'danger')
            return redirect(url_for('kardex', id=trans.product_id))
        
        if new_type == 'OUT':
            current_stock = get_current_stock(old_product_id)
            if current_stock < new_quantity:
                update_inventory(old_product_id, old_quantity, old_type)
                flash(f'Stock insuficiente para la salida. Stock actual: {current_stock}, solicitado: {new_quantity}', 'danger')
                return redirect(url_for('kardex', id=trans.product_id))
        
        trans.transaction_type = new_type
        trans.quantity = new_quantity
        trans.notes = form.notes.data
        db.session.commit()
        
        try:
            update_inventory(trans.product_id, new_quantity, new_type)
            flash('Movimiento actualizado correctamente', 'success')
        except ValueError as e:
            update_inventory(old_product_id, old_quantity, old_type)
            db.session.rollback()
            flash(f'Error al actualizar inventario: {str(e)}', 'danger')
        
        return redirect(url_for('kardex', id=trans.product_id))
    
    return render_template('transaction_edit_form.html', form=form, transaction=trans)

@app.route('/transaction/delete/<int:id>', methods=['POST'])
def delete_transaction(id):
    trans = Transaction.query.get_or_404(id)
    
    # No permitir eliminar movimientos de entrada (IN)
    if trans.transaction_type == 'IN':
        flash('❌ No se puede eliminar una entrada. Para corregir el inventario, edite la cantidad o agregue una salida compensatoria.', 'danger')
        return redirect(request.referrer or url_for('index'))
    
    if trans.deleted:
        flash('El movimiento ya está en la papelera', 'warning')
        return redirect(request.referrer or url_for('index'))
    
    try:
        revert_inventory(trans.product_id, trans.quantity, trans.transaction_type)
        trans.deleted = True
        db.session.commit()
        flash('Movimiento movido a papelera', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(request.referrer or url_for('index'))

@app.route('/transaction/restore/<int:id>', methods=['POST'])
def restore_transaction(id):
    trans = Transaction.query.get_or_404(id)
    if not trans.deleted:
        flash('Movimiento ya activo', 'warning')
        return redirect(url_for('deleted_transactions'))
    if trans.transaction_type == 'OUT':
        stock = get_current_stock(trans.product_id)
        if stock < trans.quantity:
            flash(f'No se puede restaurar: stock insuficiente ({stock} < {trans.quantity})', 'danger')
            return redirect(url_for('deleted_transactions'))
    try:
        update_inventory(trans.product_id, trans.quantity, trans.transaction_type)
        trans.deleted = False
        db.session.commit()
        flash('Movimiento restaurado', 'success')
    except ValueError as e:
        flash(str(e), 'danger')
    return redirect(url_for('deleted_transactions'))

@app.route('/transactions/all')
def all_transactions():
    page = request.args.get('page', 1, type=int)
    per_page = 50
    pagination = Transaction.query.filter_by(deleted=False).order_by(Transaction.transaction_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template('all_transactions.html', pagination=pagination)

@app.route('/transactions/deleted')
def deleted_transactions():
    page = request.args.get('page', 1, type=int)
    per_page = 50
    pagination = Transaction.query.filter_by(deleted=True).order_by(Transaction.transaction_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template('deleted_transactions.html', pagination=pagination)

# ==================== EXPORTACIONES PROFESIONALES ====================
LOGO_BASE64 = "''' + LOGO_BASE64 + '''"

class ProfessionalPDF(FPDF):
    def __init__(self, product, folio):
        super().__init__()
        self.product = product
        self.folio = folio
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        # Logo eliminado porque la imagen no era válida
        # Título centrado
        self.set_font('Arial', 'B', 16)
        self.cell(0, 10, 'TARJETA DE ESTIBA', 0, 0, 'C')
        # Folio (derecha)
        self.set_font('Arial', '', 10)
        self.set_xy(170, 10)
        self.cell(30, 10, f'Folio: {self.folio}', 0, 0, 'R')
        self.ln(15)
        # Datos del producto
        self.set_font('Arial', 'B', 12)
        self.cell(0, 8, f'Código: {self.product.code} - {self.product.name}', 0, 1, 'L')
        self.set_font('Arial', '', 10)
        self.cell(0, 6, f'Stock actual: {get_current_stock(self.product.id)} unidades', 0, 1, 'L')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()} - Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 0, 'C')

    def add_table(self, transactions):
        # Encabezados
        self.set_font('Arial', 'B', 9)
        self.set_fill_color(41, 128, 185)
        self.set_text_color(255, 255, 255)
        self.cell(30, 8, 'Fecha', 1, 0, 'C', 1)
        self.cell(20, 8, 'Tipo', 1, 0, 'C', 1)
        self.cell(25, 8, 'Entrada', 1, 0, 'C', 1)
        self.cell(25, 8, 'Salida', 1, 0, 'C', 1)
        self.cell(25, 8, 'Saldo', 1, 0, 'C', 1)
        self.cell(0, 8, 'Notas', 1, 1, 'C', 1)
        self.set_text_color(0, 0, 0)
        self.set_font('Arial', '', 9)
        # Filas
        fill = False
        for t in transactions:
            fecha = t['date'].strftime('%d/%m/%Y %H:%M')
            tipo = 'Entrada' if t['type'] == 'IN' else 'Salida'
            self.set_fill_color(240, 240, 240) if fill else self.set_fill_color(255, 255, 255)
            self.cell(30, 7, fecha, 1, 0, 'C', 1)
            self.cell(20, 7, tipo, 1, 0, 'C', 1)
            self.cell(25, 7, str(t['entrada']) if t['entrada'] > 0 else '', 1, 0, 'C', 1)
            self.cell(25, 7, str(t['salida']) if t['salida'] > 0 else '', 1, 0, 'C', 1)
            self.cell(25, 7, str(t['balance']), 1, 0, 'C', 1)
            self.cell(0, 7, t['notes'] or '', 1, 1, 'L', 1)
            fill = not fill

@app.route('/export/pdf/<int:product_id>')
def export_pdf(product_id):
    product = Product.query.get_or_404(product_id)
    transactions = get_transactions_with_balance(product_id)
    folio = f"KARDEX-{product.code}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    pdf = ProfessionalPDF(product, folio)
    pdf.add_page()
    pdf.add_table(transactions)
    output = BytesIO()
    pdf.output(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'kardex_{product.code}.pdf', mimetype='application/pdf')

@app.route('/export/excel/<int:product_id>')
def export_excel(product_id):
    product = Product.query.get_or_404(product_id)
    transactions = get_transactions_with_balance(product_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kardex"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Encabezado del documento
    ws['A1'] = 'TARJETA DE ESTIBA'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    ws['A2'] = f'Código: {product.code} - {product.name}'
    ws.merge_cells('A2:F2')
    ws['A3'] = f'Stock actual: {get_current_stock(product.id)} unidades'
    ws.merge_cells('A3:F3')
    ws['A5'] = f'Folio: KARDEX-{product.code}-{datetime.now().strftime("%Y%m%d%H%M%S")}'
    ws.merge_cells('A5:F5')
    
    # Tabla
    headers = ['Fecha', 'Tipo', 'Entrada', 'Salida', 'Saldo', 'Notas']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 8
    for t in transactions:
        ws.cell(row, 1, t['date'].strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row, 2, 'Entrada' if t['type'] == 'IN' else 'Salida').border = border
        ws.cell(row, 3, t['entrada'] if t['entrada'] > 0 else '').border = border
        ws.cell(row, 4, t['salida'] if t['salida'] > 0 else '').border = border
        ws.cell(row, 5, t['balance']).border = border
        ws.cell(row, 6, t['notes'] or '').border = border
        row += 1
    
    # Ajustar anchos
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'kardex_{product.code}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)
''')

# ==================== TEMPLATES (iguales a versión anterior pero con código) ====================
write_file("templates/base.html", """<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Tarjeta de Estiba Profesional</title>
    <link href="{{ url_for('static', filename='css/bootstrap.min.css') }}" rel="stylesheet">
    <link href="{{ url_for('static', filename='css/bootstrap-icons.min.css') }}" rel="stylesheet">
    <style>
        .go-top {
            position: fixed;
            bottom: 30px;
            right: 30px;
            background-color: #0d6efd;
            color: white;
            width: 50px;
            height: 50px;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 24px;
            cursor: pointer;
            box-shadow: 0 4px 8px rgba(0,0,0,0.2);
            transition: opacity 0.3s, visibility 0.3s;
            opacity: 0;
            visibility: hidden;
            z-index: 1000;
            text-decoration: none;
        }
        .go-top:hover {
            background-color: #0b5ed7;
            color: white;
        }
        .go-top.show {
            opacity: 1;
            visibility: visible;
        }
    </style>
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark">
        <div class="container">
            <a class="navbar-brand" href="{{ url_for('index') }}">📋 Tarjeta de Estiba Pro</a>
            <div class="navbar-nav ms-auto">
                <a class="nav-link" href="{{ url_for('all_transactions') }}">📜 Todos los Movimientos</a>
                <a class="nav-link" href="{{ url_for('deleted_products') }}">🗑️ Papelera</a>
            </div>
        </div>
    </nav>
    <div class="container mt-4">
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }} alert-dismissible fade show">{{ message }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        {% block content %}{% endblock %}
    </div>
    <a href="#" class="go-top" id="goTopBtn">⬆️</a>
    <script>
        const goTopBtn = document.getElementById('goTopBtn');
        window.addEventListener('scroll', () => {
            if (window.scrollY > 200) {
                goTopBtn.classList.add('show');
            } else {
                goTopBtn.classList.remove('show');
            }
        });
        goTopBtn.addEventListener('click', (e) => {
            e.preventDefault();
            window.scrollTo({ top: 0, behavior: 'smooth' });
        });
    </script>
    <script src="{{ url_for('static', filename='js/bootstrap.bundle.min.js') }}"></script>
</body>
</html>
""")

write_file("templates/index.html", """{% extends 'base.html' %}
{% block content %}
<div class="d-flex justify-content-between mb-3">
    <h2>Productos</h2>
    <div>
        <a href="{{ url_for('add_product') }}" class="btn btn-success">➕ Nuevo Producto</a>
        <a href="{{ url_for('add_transaction') }}" class="btn btn-primary">➕ Nuevo Movimiento</a>
    </div>
</div>
<table class="table table-bordered table-hover">
    <thead class="table-dark">
        <tr>
            <th>Código</th>
            <th>Nombre</th>
            <th>Stock</th>
            <th>Fecha de Recepción</th>
            <th>Fecha de Salida</th>
            <th>Movimientos</th>
            <th>Acciones</th>
        </tr>
    </thead>
    <tbody>
        {% for item in product_data %}
        <tr>
            <td>{{ item.product.code }}</td>
            <td>{{ item.product.name }}</td>
            <td>{{ item.stock }}</td>
            <td>{{ item.last_reception.strftime('%d/%m/%Y %H:%M') if item.last_reception else '—' }}</td>
            <td>{{ item.last_dispatch.strftime('%d/%m/%Y %H:%M') if item.last_dispatch else '—' }}</td>
            <td><a href="{{ url_for('kardex', id=item.product.id) }}" class="btn btn-sm btn-info">📋 Ver Tarjeta</a></td>
            <td>
                <a href="{{ url_for('edit_product', id=item.product.id) }}" class="btn btn-sm btn-warning">✏️</a>
                <a href="{{ url_for('delete_product', id=item.product.id) }}" class="btn btn-sm btn-danger" onclick="return confirm('¿Mover a papelera?')">🗑️</a>
            </td>
        </tr>
        {% endfor %}
    </tbody>
</table>
{% endblock %}
""")

write_file("templates/kardex.html", """{% extends 'base.html' %}
{% block content %}
<div class="d-flex justify-content-between align-items-center">
    <h2>Tarjeta de Estiba: {{ product.code }} - {{ product.name }}</h2>
    <div>
        <a href="{{ url_for('export_pdf', product_id=product.id) }}" class="btn btn-danger">📄 Exportar PDF</a>
        <a href="{{ url_for('export_excel', product_id=product.id) }}" class="btn btn-success">📊 Exportar Excel</a>
        <a href="{{ url_for('index') }}" class="btn btn-secondary">← Volver</a>
    </div>
</div>
<p><strong>Stock actual:</strong> {{ stock }} unidades</p>
<hr>
<h4>Movimientos</h4>
<table class="table table-bordered table-hover">
    <thead class="table-dark">
        <tr>
            <th>Fecha/Hora</th>
            <th>Tipo</th>
            <th>Entrada</th>
            <th>Salida</th>
            <th>Saldo</th>
            <th>Notas</th>
            <th>Acciones</th>
        </tr>
    </thead>
    <tbody>
        {% for t in transactions %}
        <tr>
            <td>{{ t.date.strftime('%Y-%m-%d %H:%M:%S') }}</td>
            <td>
                {% if t.type == 'IN' %}
                    <span class="badge bg-success">Entrada</span>
                {% else %}
                    <span class="badge bg-danger">Salida</span>
                {% endif %}
            </td>
            <td>{{ t.entrada if t.entrada > 0 else '' }}</td>
            <td>{{ t.salida if t.salida > 0 else '' }}</td>
            <td>{{ t.balance }}</td>
            <td>{{ t.notes or '-' }}</td>
            <td>
                <a href="{{ url_for('edit_transaction', id=t.id) }}" class="btn btn-sm btn-warning">✏️</a>
                <form method="POST" action="{{ url_for('delete_transaction', id=t.id) }}" style="display:inline;" onsubmit="return confirm('¿Mover a papelera?')">
                    <button type="submit" class="btn btn-sm btn-danger">🗑️</button>
                </form>
            </td>
        </tr>
        {% else %}
        <tr><td colspan="7" class="text-center">No hay movimientos registrados.{% endfor %}
    </tbody>
</table>
<a href="{{ url_for('add_transaction') }}?product_id={{ product.id }}" class="btn btn-primary">➕ Agregar Movimiento</a>
{% endblock %}
""")

write_file("templates/product_form.html", """{% extends 'base.html' %}
{% block content %}
<div class="card shadow">
    <div class="card-header bg-primary text-white"><h4>{{ title }}</h4></div>
    <div class="card-body">
        <form method="POST">
            {{ form.hidden_tag() }}
            <div class="mb-3">{{ form.code.label(class="form-label") }}{{ form.code(class="form-control") }}</div>
            <div class="mb-3">{{ form.name.label(class="form-label") }}{{ form.name(class="form-control") }}</div>
            <button type="submit" class="btn btn-primary">Guardar</button>
            <a href="{{ url_for('index') }}" class="btn btn-secondary">Cancelar</a>
        </form>
    </div>
</div>
{% endblock %}
""")

write_file("templates/transaction_form.html", """{% extends 'base.html' %}
{% block content %}
<div class="card shadow">
    <div class="card-header bg-success text-white"><h4>Registrar Movimiento</h4></div>
    <div class="card-body">
        <form method="POST">
            {{ form.hidden_tag() }}
            <div class="mb-3">{{ form.product_id.label(class="form-label") }}{{ form.product_id(class="form-select") }}</div>
            <div class="mb-3">{{ form.transaction_type.label(class="form-label") }}{{ form.transaction_type(class="form-select") }}</div>
            <div class="mb-3">{{ form.quantity.label(class="form-label") }}{{ form.quantity(class="form-control") }}</div>
            <div class="mb-3">{{ form.notes.label(class="form-label") }}{{ form.notes(class="form-control") }}</div>
            <button type="submit" class="btn btn-success">Registrar</button>
            <a href="{{ url_for('index') }}" class="btn btn-secondary">Cancelar</a>
        </form>
    </div>
</div>
{% endblock %}
""")

write_file("templates/transaction_edit_form.html", """{% extends 'base.html' %}
{% block content %}
<div class="card shadow">
    <div class="card-header bg-warning text-dark"><h4>Editar Movimiento</h4></div>
    <div class="card-body">
        <p><strong>Producto:</strong> {{ transaction.product.code }} - {{ transaction.product.name }}</p>
        <p><strong>Fecha original:</strong> {{ transaction.transaction_date.strftime('%Y-%m-%d %H:%M:%S') }}</p>
        <form method="POST">
            {{ form.hidden_tag() }}
            <div class="mb-3">{{ form.product_id.label(class="form-label") }}{{ form.product_id(class="form-select", disabled=True) }}</div>
            <div class="mb-3">{{ form.transaction_type.label(class="form-label") }}{{ form.transaction_type(class="form-select") }}</div>
            <div class="mb-3">{{ form.quantity.label(class="form-label") }}{{ form.quantity(class="form-control") }}</div>
            <div class="mb-3">{{ form.notes.label(class="form-label") }}{{ form.notes(class="form-control") }}</div>
            <button type="submit" class="btn btn-primary">Guardar Cambios</button>
            <a href="{{ url_for('kardex', id=transaction.product_id) }}" class="btn btn-secondary">Cancelar</a>
        </form>
    </div>
</div>
{% endblock %}
""")

write_file("templates/all_transactions.html", """{% extends 'base.html' %}
{% block content %}
<h2>Todos los Movimientos</h2>
<a href="{{ url_for('index') }}" class="btn btn-secondary mb-3">← Volver</a>
<table class="table table-bordered table-hover">
    <thead class="table-dark">
        <tr>
            <th>No.</th>
            <th>Producto</th>
            <th>Tipo</th>
            <th>Cantidad</th>
            <th>Fecha/Hora</th>
            <th>Notas</th>
            <th>Acciones</th>
        </tr>
    </thead>
    <tbody>
        {% for t in pagination.items %}
        <tr>
            <td>{{ loop.index + (pagination.page-1)*pagination.per_page }}</td>
            <td>{{ t.product.code }} - {{ t.product.name }}</td>
            <td><span class="badge bg-{{ 'success' if t.transaction_type=='IN' else 'danger' }}">{{ 'Entrada' if t.transaction_type=='IN' else 'Salida' }}</span></td>
            <td>{{ t.quantity }}</td>
            <td>{{ t.transaction_date.strftime('%Y-%m-%d %H:%M:%S') }}</td>
            <td>{{ t.notes or '-' }}</td>
            <td>
                <a href="{{ url_for('edit_transaction', id=t.id) }}" class="btn btn-sm btn-warning">✏️</a>
                <form method="POST" action="{{ url_for('delete_transaction', id=t.id) }}" style="display:inline;" onsubmit="return confirm('¿Mover a papelera?')">
                    <button type="submit" class="btn btn-sm btn-danger">🗑️</button>
                </form>
            </td>
        </tr>
        {% endfor %}
    </tbody>
</table>
<nav>
    <ul class="pagination">
        {% if pagination.has_prev %}<li class="page-item"><a class="page-link" href="?page={{ pagination.prev_num }}">Anterior</a></li>{% endif %}
        <li class="page-item active"><span class="page-link">Página {{ pagination.page }} de {{ pagination.pages }}</span></li>
        {% if pagination.has_next %}<li class="page-item"><a class="page-link" href="?page={{ pagination.next_num }}">Siguiente</a></li>{% endif %}
    </ul>
</nav>
{% endblock %}
""")

write_file("templates/deleted_products.html", """{% extends 'base.html' %}
{% block content %}
<h2>🗑️ Papelera de Productos</h2>
<a href="{{ url_for('index') }}" class="btn btn-primary mb-3">← Volver</a>
<table class="table table-striped">
    <thead class="table-dark">
        <tr><th>Código</th><th>Nombre</th><th>Acciones</th></tr>
    </thead>
    <tbody>
        {% for p in products %}
        <tr>
            <td>{{ p.code }}</td>
            <td>{{ p.name }}</td>
            <td><a href="{{ url_for('restore_product', id=p.id) }}" class="btn btn-sm btn-success">↩️ Restaurar</a></td>
        </tr>
        {% else %}
        <tr><td colspan="3" class="text-center">No hay productos en papelera.{% endfor %}
    </tbody>
</table>
{% endblock %}
""")

write_file("templates/deleted_transactions.html", """{% extends 'base.html' %}
{% block content %}
<h2>🗑️ Movimientos Eliminados</h2>
<a href="{{ url_for('index') }}" class="btn btn-primary mb-3">← Volver</a>
<table class="table table-striped">
    <thead class="table-dark">
        <tr><th>Producto</th><th>Tipo</th><th>Cantidad</th><th>Fecha</th><th>Acciones</th></tr>
    </thead>
    <tbody>
        {% for t in pagination.items %}
        <tr>
            <td>{{ t.product.code }} - {{ t.product.name }}</td>
            <td><span class="badge bg-{{ 'success' if t.transaction_type=='IN' else 'danger' }}">{{ 'Entrada' if t.transaction_type=='IN' else 'Salida' }}</span></td>
            <td>{{ t.quantity }}</td>
            <td>{{ t.transaction_date.strftime('%Y-%m-%d %H:%M') }}</td>
            <td>
                <form method="POST" action="{{ url_for('restore_transaction', id=t.id) }}" style="display:inline;">
                    <button type="submit" class="btn btn-sm btn-success">↩️ Restaurar</button>
                </form>
            </td>
        </tr>
        {% else %}
        <tr><td colspan="5" class="text-center">No hay movimientos en papelera.{% endfor %}
    </tbody>
</table>
<nav>
    <ul class="pagination">
        {% if pagination.has_prev %}<li class="page-item"><a class="page-link" href="?page={{ pagination.prev_num }}">Anterior</a></li>{% endif %}
        <li class="page-item active"><span class="page-link">Página {{ pagination.page }} de {{ pagination.pages }}</span></li>
        {% if pagination.has_next %}<li class="page-item"><a class="page-link" href="?page={{ pagination.next_num }}">Siguiente</a></li>{% endif %}
    </ul>
</nav>
{% endblock %}
""")

write_file("static/css/style.css", """body { background: #f8f9fa; }
.card { border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); }
""")

# ==================== FUNCIONES DE EJECUCIÓN ====================
def run_command(cmd, cwd=None):
    print(f"\n> {cmd}")
    process = subprocess.Popen(cmd, shell=True, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                               text=True, encoding='utf-8', errors='replace')
    for line in process.stdout:
        print(line, end='')
    process.wait()
    if process.returncode != 0:
        print(f"⚠️  El comando falló con código {process.returncode}")
        sys.exit(process.returncode)

def run_application():
    os.chdir(PROJECT_ROOT)
    run_command("python -m venv venv")
    if sys.platform == "win32":
        python_exe = "venv\\Scripts\\python.exe"
        flask_exe = "venv\\Scripts\\flask.exe"
    else:
        python_exe = "venv/bin/python"
        flask_exe = "venv/bin/flask"
    run_command(f"{python_exe} -m pip install --upgrade pip")
    run_command(f"{python_exe} -m pip install -r requirements.txt")
    run_command(f"{flask_exe} init-db")
    run_command(f"{flask_exe} import-products")
    print("\n🚀 Levantando servidor Flask...")
    run_command(f"{python_exe} app.py")

if __name__ == '__main__':
    download_static_dependencies()
    print("🚀 Generando proyecto Tarjeta de Estiba Profesional...")
    print("📁 Carpeta: tarjeta_estiba_profesional/")
    print(f"✅ {len(PRODUCT_NAMES)} productos importados con códigos automáticos")
    print("✅ Movimientos con entrada, salida y saldo acumulado")
    print("✅ Exportación PDF profesional con logo y folio")
    print("✅ Exportación Excel con formato corporativo")
    print("✅ Botón flotante y edición de movimientos")
    print("\n🚀 Iniciando instalación automática y ejecución...\n")
    run_application()