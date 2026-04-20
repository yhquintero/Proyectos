#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🚀 GENERADOR DE TARJETA DE ESTIBA (KARDEX) - SIN VALORES MONETARIOS
Solo productos, cantidades y movimientos. Exportación a PDF y Excel.
"""

import urllib.request
import ssl
import os
import re
import json
import subprocess
import sys
from pathlib import Path
from datetime import datetime

# # ==================== RAW_DATA (solo nombres, sin precios) ====================
RAW_NAMES = """
Aceite
Aceite 1 Litro 1400
Bombón
Botonetas
Café Ziva
Cerveza 3 Caballos 210
Cerveza 3 Caballos 230
Cerveza Breda
Cerveza Cristal 280
Cerveza Mayabe 230
Cerveza Mayabe 280
Cerveza Parranda 210
Cerveza Parranda 230
Cerveza Presidente 210
Cerveza Presidente 250
Cerveza Shekels 220
Cerveza Shekels 250
Cerveza Unlager 220
Chambelona
Chicle
Chupa (60)
Cigarros H Upman
Cigarros Popular Rojo
Coditos 500g
Detergente 1kg
Detergente 500g 500
Embutido Jamón
Energizante
Galleta Craker
Galleta Lamore
Galleta Minees 70
Galleta Pygmi
Galleta Salrica
Galletas de sal
Galletas Luanezco
Galletas My Bit
Habana Club Añejo Lt
Jabón
Jabón 200
Jugos
Malta Guajira 330 ml
Malta Vitali
Menta Plus
Mybar
Pasta Dental
Pelly
Peter Cars 100
Picadillo
Pollo Ahumado 300
Pure de Tomate
Refresco Cola (1.5L)
Refresco Cola (lata) 250
Refresco Cotel paquete
Refresco Guanabana (paquete)
Refresco Limon (Lata) 250
Refresco Mandarina (paquete)
Refresco Melocotón (paquete)
Refresco Naranja (1.5L)
Refresco Naranja (Lata) 250
Refresco Pepsi
Refresco Piña Colada (paquete)
Refresco Pomo 250ml
Rona granel 200 ml
Rona Arocha 260
Rona Arocha 500 ml
Ron cajio venta granel 200 ml
Ron Guayabita
Ron Nucay (granel) 350 ml
Ron Old Partner
Ron Rivera
Ron Vega del Rio
Sal Común
Salchicha 480 g
Sazón Comino
Sazón Guama
Sopa
Sorbeto Piña Colada (paquete)
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Spaguetti
Togo Peanut
Vinagre
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Sorbeto Fist
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Wisky 375 ml
Sorbeto First
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Vodka Pepe
Vodka Peanut
Vodka Mágico
"""

def parse_names(raw_text):
    lines = [line.strip() for line in raw_text.strip().split('\n') if line.strip()]
    # eliminar duplicados manteniendo orden
    seen = set()
    unique = []
    for name in lines:
        if name not in seen:
            seen.add(name)
            unique.append(name)
    return unique

PRODUCT_NAMES = parse_names(RAW_NAMES)
PROJECT_ROOT = Path("tarjeta_estiba_sin_valores")

def write_file(relative_path: str, content: str):
    path = PROJECT_ROOT / relative_path
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding='utf-8')
    print(f"✅ Creado: {path}")

def download_file(url, dest_path):
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        ctx = ssl._create_unverified_context()
        with urllib.request.urlopen(url, context=ctx) as response, open(dest_path, 'wb') as out_file:
            out_file.write(response.read())
        print(f"📥 Descargado: {dest_path}")
    except Exception as e:
        print(f"❌ Error descargando {url}: {e}")

def download_static_dependencies():
    static_dir = PROJECT_ROOT / "static"
    bootstrap_css = static_dir / "css" / "bootstrap.min.css"
    if not bootstrap_css.exists():
        download_file("https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css", bootstrap_css)
    bootstrap_js = static_dir / "js" / "bootstrap.bundle.min.js"
    if not bootstrap_js.exists():
        download_file("https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js", bootstrap_js)
    icons_css = static_dir / "css" / "bootstrap-icons.min.css"
    if not icons_css.exists():
        download_file("https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css", icons_css)
    icons_font_dir = static_dir / "css" / "fonts"
    icons_font_files = [
        "https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/fonts/bootstrap-icons.woff2",
        "https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/fonts/bootstrap-icons.woff"
    ]
    for font_url in icons_font_files:
        font_name = font_url.split("/")[-1]
        font_path = icons_font_dir / font_name
        if not font_path.exists():
            download_file(font_url, font_path)
    # No se necesita Chart.js, pero se puede omitir

# # ==================== ARCHIVOS DEL PROYECTO ====================
write_file("requirements.txt", """Flask==3.0.3
Flask-SQLAlchemy==3.1.1
Flask-WTF==1.2.1
WTForms==3.1.2
python-dotenv==1.0.1
psycopg2-binary==2.9.9
click==8.1.7
fpdf2==2.7.9
openpyxl==3.1.2
""")

write_file(".env", """DATABASE_URL=sqlite:///kardex.db
SECRET_KEY=clave-segura-para-desarrollo
FLASK_ENV=development
""")

write_file("README.md", """# Tarjeta de Estiba - Sin Valores Monetarios

Sistema de control de inventario por producto, solo cantidades y movimientos.

## Características
- Gestión de productos (CRUD)
- Registro de entradas y salidas con cantidad
- **Sin precios ni costos** (solo cantidades)
- Tarjeta de estiba por producto con exportación a PDF y Excel
- Listado de todos los movimientos
- Papelera de productos y movimientos
- Fecha automática en cada movimiento

## Instalación y ejecución
```bash
cd tarjeta_estiba_sin_valores
python -m venv venv
source venv/bin/activate  # o venv\\Scripts\\activate
pip install -r requirements.txt
flask init-db
flask import-products
flask run
""")

write_file("models.py", """from flask_sqlalchemy import SQLAlchemy
from datetime import datetime

db = SQLAlchemy()

class Product(db.Model):
tablename = 'products'
id = db.Column(db.Integer, primary_key=True)
name = db.Column(db.String(200), unique=True, nullable=False)
deleted = db.Column(db.Boolean, default=False)
created_at = db.Column(db.DateTime, default=datetime.utcnow)
inventory = db.relationship('Inventory', backref='product', uselist=False)
transactions = db.relationship('Transaction', backref='product', lazy=True)

class Inventory(db.Model):
tablename = 'inventories'
product_id = db.Column(db.Integer, db.ForeignKey('products.id'), primary_key=True)
quantity = db.Column(db.Integer, default=0, nullable=False)
last_updated = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Transaction(db.Model):
tablename = 'transactions'
id = db.Column(db.Integer, primary_key=True)
product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
transaction_type = db.Column(db.String(3), nullable=False) # 'IN' o 'OUT'
quantity = db.Column(db.Integer, nullable=False)
transaction_date = db.Column(db.DateTime, default=datetime.utcnow)
notes = db.Column(db.String(300))
deleted = db.Column(db.Boolean, default=False)
""")

write_file("forms.py", """from flask_wtf import FlaskForm
from wtforms import StringField, IntegerField, SelectField, TextAreaField
from wtforms.validators import DataRequired, NumberRange, ValidationError
from models import Product

class ProductForm(FlaskForm):
name = StringField('Nombre del Producto', validators=[DataRequired()])

def validate_name(self, field):
product = Product.query.filter_by(name=field.data, deleted=False).first()
if product and product.id != getattr(self, '_product_id', None):
raise ValidationError('Ya existe un producto con este nombre.')

class TransactionForm(FlaskForm):
product_id = SelectField('Producto', coerce=int, validators=[DataRequired()])
transaction_type = SelectField('Tipo', choices=[('IN', 'Entrada'), ('OUT', 'Salida')], validators=[DataRequired()])
quantity = IntegerField('Cantidad', validators=[DataRequired(), NumberRange(min=1)])
notes = TextAreaField('Notas')
""")

write_file("utils.py", """from datetime import datetime
from models import db, Transaction, Product, Inventory

def update_inventory(product_id, quantity, transaction_type):
inv = Inventory.query.get(product_id)
if not inv:
inv = Inventory(product_id=product_id, quantity=0)
db.session.add(inv)
if transaction_type == 'IN':
inv.quantity += quantity
else:
if inv.quantity < quantity:
raise ValueError(f"Stock insuficiente: {inv.quantity} < {quantity}")
inv.quantity -= quantity
db.session.commit()

def revert_inventory(product_id, old_quantity, old_type):
inv = Inventory.query.get(product_id)
if not inv:
inv = Inventory(product_id=product_id, quantity=0)
db.session.add(inv)
if old_type == 'IN':
inv.quantity -= old_quantity
else:
inv.quantity += old_quantity
if inv.quantity < 0:
inv.quantity = 0
db.session.commit()

def get_current_stock(product_id):
inv = Inventory.query.get(product_id)
return inv.quantity if inv else 0
""")

write_file("app.py", '''
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from models import db, Product, Inventory, Transaction
from forms import ProductForm, TransactionForm
from utils import update_inventory, revert_inventory, get_current_stock
from datetime import datetime
import click
import os
import re
from io import BytesIO
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment
from dotenv import load_dotenv

load_dotenv()

app = Flask(name)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'clave-segura-por-defecto')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///kardex.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# ==================== DATOS RAW (solo nombres) ====================
RAW_NAMES = """
Aceite
Aceite 1 Litro 1400
Bombón
Botonetas
Café Ziva
Cerveza 3 Caballos 210
Cerveza 3 Caballos 230
Cerveza Breda
Cerveza Cristal 280
Cerveza Mayabe 230
Cerveza Mayabe 280
Cerveza Parranda 210
Cerveza Parranda 230
Cerveza Presidente 210
Cerveza Presidente 250
Cerveza Shekels 220
Cerveza Shekels 250
Cerveza Unlager 220
Chambelona
Chicle
Chupa (60)
Cigarros H Upman
Cigarros Popular Rojo
Coditos 500g
Detergente 1kg
Detergente 500g 500
Embutido Jamón
Energizante
Galleta Craker
Galleta Lamore
Galleta Minees 70
Galleta Pygmi
Galleta Salrica
Galletas de sal
Galletas Luanezco
Galletas My Bit
Habana Club Añejo Lt
Jabón
Jabón 200
Jugos
Malta Guajira 330 ml
Malta Vitali
Menta Plus
Mybar
Pasta Dental
Pelly
Peter Cars 100
Picadillo
Pollo Ahumado 300
Pure de Tomate
Refresco Cola (1.5L)
Refresco Cola (lata) 250
Refresco Cotel paquete
Refresco Guanabana (paquete)
Refresco Limon (Lata) 250
Refresco Mandarina (paquete)
Refresco Melocotón (paquete)
Refresco Naranja (1.5L)
Refresco Naranja (Lata) 250
Refresco Pepsi
Refresco Piña Colada (paquete)
Refresco Pomo 250ml
Rona granel 200 ml
Rona Arocha 260
Rona Arocha 500 ml
Ron cajio venta granel 200 ml
Ron Guayabita
Ron Nucay (granel) 350 ml
Ron Old Partner
Ron Rivera
Ron Vega del Rio
Sal Común
Salchicha 480 g
Sazón Comino
Sazón Guama
Sopa
Sorbeto Piña Colada (paquete)
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Spaguetti
Togo Peanut
Vinagre
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Sorbeto Fist
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Wisky 375 ml
Sorbeto First
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Sorbeto Nuto 250
Sorbeto Renata
Sorbeto Togo
Sorbeto Pepe
Sorbeto Peanut
Sorbeto Mege
Vip Max
Vodka Mágico
Vodka NA ZOOROVIE 40 Grado
Vodka Royal
Vodka Tabarish
Vodka Pepe
Vodka Peanut
Vodka Mágico
"""

def parse_names(raw_text):
lines = [line.strip() for line in raw_text.strip().split('\n') if line.strip()]
seen = set()
unique = []
for name in lines:
if name not in seen:
seen.add(name)
unique.append(name)
return unique

# ==================== COMANDOS FLASK ====================
@app.cli.command('init-db')
def init_db_command():
db.create_all()
click.echo('[OK] Base de datos inicializada.')

@app.cli.command('import-products')
def import_products():
names = parse_names(RAW_NAMES)
count = 0
for name in names:
if not Product.query.filter_by(name=name, deleted=False).first():
prod = Product(name=name)
db.session.add(prod)
db.session.flush()
db.session.add(Inventory(product_id=prod.id, quantity=0))
count += 1
db.session.commit()
click.echo(f'[OK] Se importaron {count} productos.')

# ==================== RUTAS ====================
@app.route('/')
def index():
products = Product.query.filter_by(deleted=False).all()
return render_template('index.html', products=products)

@app.route('/product/<int:id>/kardex')
def kardex(id):
product = Product.query.get_or_404(id)
transactions = Transaction.query.filter_by(product_id=id, deleted=False).order_by(Transaction.transaction_date.desc()).all()
stock = get_current_stock(id)
return render_template('kardex.html', product=product, transactions=transactions, stock=stock)

@app.route('/product/add', methods=['GET', 'POST'])
def add_product():
form = ProductForm()
if form.validate_on_submit():
product = Product(name=form.name.data)
db.session.add(product)
db.session.flush()
db.session.add(Inventory(product_id=product.id, quantity=0))
db.session.commit()
flash('Producto agregado correctamente', 'success')
return redirect(url_for('index'))
return render_template('product_form.html', form=form, title='Nuevo Producto')

@app.route('/product/edit/<int:id>', methods=['GET', 'POST'])
def edit_product(id):
product = Product.query.get_or_404(id)
form = ProductForm(obj=product)
form._product_id = id
if form.validate_on_submit():
product.name = form.name.data
db.session.commit()
flash('Producto actualizado', 'success')
return redirect(url_for('index'))
return render_template('product_form.html', form=form, title='Editar Producto')

@app.route('/product/delete/<int:id>')
def delete_product(id):
product = Product.query.get_or_404(id)
product.deleted = True
db.session.commit()
flash('Producto movido a papelera', 'warning')
return redirect(url_for('index'))

@app.route('/product/restore/<int:id>')
def restore_product(id):
product = Product.query.get_or_404(id)
product.deleted = False
db.session.commit()
flash('Producto restaurado', 'success')
return redirect(url_for('deleted_products'))

@app.route('/products/deleted')
def deleted_products():
products = Product.query.filter_by(deleted=True).all()
return render_template('deleted_products.html', products=products)

@app.route('/transaction/add', methods=['GET', 'POST'])
def add_transaction():
form = TransactionForm()
form.product_id.choices = [(p.id, p.name) for p in Product.query.filter_by(deleted=False).order_by(Product.name).all()]
if form.validate_on_submit():
product = Product.query.get(form.product_id.data)
now = datetime.now()
if form.transaction_type.data == 'OUT':
stock = get_current_stock(product.id)
if stock < form.quantity.data:
flash(f'Stock insuficiente. Disponible: {stock}, solicitado: {form.quantity.data}', 'danger')
return redirect(url_for('add_transaction'))

trans = Transaction(
product_id=product.id,
transaction_type=form.transaction_type.data,
quantity=form.quantity.data,
transaction_date=now,
notes=form.notes.data,
deleted=False
)
db.session.add(trans)
try:
update_inventory(product.id, form.quantity.data, form.transaction_type.data)
db.session.commit()
flash('Movimiento registrado', 'success')
return redirect(url_for('kardex', id=product.id))
except ValueError as e:
db.session.rollback()
flash(str(e), 'danger')
return render_template('transaction_form.html', form=form)

@app.route('/transaction/delete/<int:id>', methods=['POST'])
def delete_transaction(id):
trans = Transaction.query.get_or_404(id)
if trans.deleted:
flash('Movimiento ya está en papelera', 'warning')
return redirect(request.referrer or url_for('index'))
try:
revert_inventory(trans.product_id, trans.quantity, trans.transaction_type)
trans.deleted = True
db.session.commit()
flash('Movimiento movido a papelera', 'success')
except Exception as e:
db.session.rollback()
flash(f'Error: {str(e)}', 'danger')
return redirect(request.referrer or url_for('index'))

@app.route('/transaction/restore/<int:id>', methods=['POST'])
def restore_transaction(id):
trans = Transaction.query.get_or_404(id)
if not trans.deleted:
flash('Movimiento ya activo', 'warning')
return redirect(url_for('deleted_transactions'))
if trans.transaction_type == 'OUT':
stock = get_current_stock(trans.product_id)
if stock < trans.quantity:
flash(f'No se puede restaurar: stock insuficiente ({stock} < {trans.quantity})', 'danger')
return redirect(url_for('deleted_transactions'))
try:
update_inventory(trans.product_id, trans.quantity, trans.transaction_type)
trans.deleted = False
db.session.commit()
flash('Movimiento restaurado', 'success')
except ValueError as e:
flash(str(e), 'danger')
return redirect(url_for('deleted_transactions'))

@app.route('/transactions/all')
def all_transactions():
page = request.args.get('page', 1, type=int)
per_page = 50
pagination = Transaction.query.filter_by(deleted=False).order_by(Transaction.transaction_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
return render_template('all_transactions.html', pagination=pagination)

@app.route('/transactions/deleted')
def deleted_transactions():
page = request.args.get('page', 1, type=int)
per_page = 50
pagination = Transaction.query.filter_by(deleted=True).order_by(Transaction.transaction_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
return render_template('deleted_transactions.html', pagination=pagination)

# ==================== EXPORTACIONES ====================
class KardexPDF(FPDF):
def header(self):
self.set_font('Arial', 'B', 12)
self.cell(0, 10, 'Tarjeta de Estiba', 0, 1, 'C')
self.ln(5)
def footer(self):
self.set_y(-15)
self.set_font('Arial', 'I', 8)
self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')

@app.route('/export/pdf/<int:product_id>')
def export_pdf(product_id):
product = Product.query.get_or_404(product_id)
transactions = Transaction.query.filter_by(product_id=product_id, deleted=False).order_by(Transaction.transaction_date).all()
stock = get_current_stock(product_id)

pdf = KardexPDF()
pdf.add_page()
pdf.set_font('Arial', 'B', 14)
pdf.cell(0, 10, f'Producto: {product.name}', 0, 1, 'L')
pdf.set_font('Arial', '', 12)
pdf.cell(0, 10, f'Stock actual: {stock} unidades', 0, 1, 'L')
pdf.ln(5)

Tabla
pdf.set_font('Arial', 'B', 10)
pdf.cell(30, 8, 'Fecha', 1, align='C')
pdf.cell(25, 8, 'Tipo', 1, align='C')
pdf.cell(25, 8, 'Cantidad', 1, align='C')
pdf.cell(0, 8, 'Notas', 1, align='C')
pdf.ln()

pdf.set_font('Arial', '', 9)
for t in transactions:
fecha = t.transaction_date.strftime('%d/%m/%Y %H:%M')
tipo = 'Entrada' if t.transaction_type == 'IN' else 'Salida'
pdf.cell(30, 8, fecha, 1)
pdf.cell(25, 8, tipo, 1, align='C')
pdf.cell(25, 8, str(t.quantity), 1, align='C')
pdf.cell(0, 8, t.notes or '-', 1)
pdf.ln()

output = BytesIO()
pdf.output(output)
output.seek(0)
return send_file(output, as_attachment=True, download_name=f'kardex_{product.name}.pdf', mimetype='application/pdf')

@app.route('/export/excel/<int:product_id>')
def export_excel(product_id):
product = Product.query.get_or_404(product_id)
transactions = Transaction.query.filter_by(product_id=product_id, deleted=False).order_by(Transaction.transaction_date).all()
stock = get_current_stock(product_id)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kardex"
ws['A1'] = f'Producto: {product.name}'
ws['A2'] = f'Stock actual: {stock} unidades'
ws['A4'] = 'Fecha'; ws['B4'] = 'Tipo'; ws['C4'] = 'Cantidad'; ws['D4'] = 'Notas'
for cell in ws['4:4']:
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal='center')

row = 5
for t in transactions:
ws.cell(row, 1, t.transaction_date.strftime('%d/%m/%Y %H:%M'))
ws.cell(row, 2, 'Entrada' if t.transaction_type == 'IN' else 'Salida')
ws.cell(row, 3, t.quantity)
ws.cell(row, 4, t.notes or '-')
row += 1

for col in ['A', 'B', 'C', 'D']:
ws.column_dimensions[col].width = 20

output = BytesIO()
wb.save(output)
output.seek(0)
return send_file(output, as_attachment=True, download_name=f'kardex_{product.name}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if name == 'main':
app.run(debug=True)
''')

# ==================== TEMPLATES ====================
write_file("templates/base.html", """<!DOCTYPE html>

<html lang="es"> <head> <meta charset="UTF-8"> <meta name="viewport" content="width=device-width, initial-scale=1.0"> <title>Tarjeta de Estiba</title> <link href="{{ url_for('static', filename='css/bootstrap.min.css') }}" rel="stylesheet"> <link href="{{ url_for('static', filename='css/bootstrap-icons.min.css') }}" rel="stylesheet"> </head> <body> <nav class="navbar navbar-expand-lg navbar-dark bg-dark"> <div class="container"> <a class="navbar-brand" href="{{ url_for('index') }}">📋 Tarjeta de Estiba</a> <div class="navbar-nav ms-auto"> <a class="nav-link" href="{{ url_for('all_transactions') }}">📜 Todos los Movimientos</a> <a class="nav-link" href="{{ url_for('deleted_products') }}">🗑️ Papelera</a> </div> </div> </nav> <div class="container mt-4"> {% with messages = get_flashed_messages(with_categories=true) %} {% if messages %} {% for category, message in messages %} <div class="alert alert-{{ category }} alert-dismissible fade show">{{ message }}</div> {% endfor %} {% endif %} {% endwith %} {% block content %}{% endblock %} </div> <script src="{{ url_for('static', filename='js/bootstrap.bundle.min.js') }}"></script> </body> </html> """)
write_file("templates/index.html", """{% extends 'base.html' %}
{% block content %}

<div class="d-flex justify-content-between mb-3"> <h2>Productos</h2> <div> <a href="{{ url_for('add_product') }}" class="btn btn-success">➕ Nuevo Producto</a> <a href="{{ url_for('add_transaction') }}" class="btn btn-primary">➕ Nuevo Movimiento</a> </div> </div> <table class="table table-bordered table-hover"> <thead class="table-dark"> <tr> <th>No.</th> <th>Nombre</th> <th>Stock</th> <th>Movimientos</th> <th>Acciones</th> </tr> </thead> <tbody> {% for p in products %} <tr> <td>{{ loop.index }}</td> <td>{{ p.name }}</td> <td>{{ p.inventory.quantity if p.inventory else 0 }}</td> <td><a href="{{ url_for('kardex', id=p.id) }}" class="btn btn-sm btn-info">📋 Ver Tarjeta</a></td> <td> <a href="{{ url_for('edit_product', id=p.id) }}" class="btn btn-sm btn-warning">✏️</a> <a href="{{ url_for('delete_product', id=p.id) }}" class="btn btn-sm btn-danger" onclick="return confirm('¿Mover a papelera?')">🗑️</a> </td> </tr> {% endfor %} </tbody> </table> {% endblock %} """)
write_file("templates/kardex.html", """{% extends 'base.html' %}
{% block content %}

<div class="d-flex justify-content-between align-items-center"> <h2>Tarjeta de Estiba: {{ product.name }}</h2> <div> <a href="{{ url_for('export_pdf', product_id=product.id) }}" class="btn btn-danger">📄 Exportar PDF</a> <a href="{{ url_for('export_excel', product_id=product.id) }}" class="btn btn-success">📊 Exportar Excel</a> <a href="{{ url_for('index') }}" class="btn btn-secondary">← Volver</a> </div> </div> <p><strong>Stock actual:</strong> {{ stock }} unidades</p> <hr> <h4>Movimientos</h4> <table class="table table-bordered table-hover"> <thead class="table-dark"> <tr> <th>Fecha/Hora</th> <th>Tipo</th> <th>Cantidad</th> <th>Notas</th> <th>Acciones</th> </tr> </thead> <tbody> {% for t in transactions %} <tr> <td>{{ t.transaction_date.strftime('%Y-%m-%d %H:%M:%S') }}</td> <td> {% if t.transaction_type == 'IN' %} <span class="badge bg-success">Entrada</span> {% else %} <span class="badge bg-danger">Salida</span> {% endif %} </td> <td>{{ t.quantity }}</td> <td>{{ t.notes or '-' }}</td> <td> <form method="POST" action="{{ url_for('delete_transaction', id=t.id) }}" style="display:inline;" onsubmit="return confirm('¿Mover a papelera?')"> <button type="submit" class="btn btn-sm btn-danger">🗑️</button> </form> </td> </tr> {% else %} <tr><td colspan="5" class="text-center">No hay movimientos registrados.{% endfor %} </tbody> </table> <a href="{{ url_for('add_transaction') }}?product_id={{ product.id }}" class="btn btn-primary">➕ Agregar Movimiento</a> {% endblock %} """)
write_file("templates/product_form.html", """{% extends 'base.html' %}
{% block content %}

<div class="card shadow"> <div class="card-header bg-primary text-white"><h4>{{ title }}</h4></div> <div class="card-body"> <form method="POST"> {{ form.hidden_tag() }} <div class="mb-3">{{ form.name.label(class="form-label") }}{{ form.name(class="form-control") }}</div> <button type="submit" class="btn btn-primary">Guardar</button> <a href="{{ url_for('index') }}" class="btn btn-secondary">Cancelar</a> </form> </div> </div> {% endblock %} """)
write_file("templates/transaction_form.html", """{% extends 'base.html' %}
{% block content %}

<div class="card shadow"> <div class="card-header bg-success text-white"><h4>Registrar Movimiento</h4></div> <div class="card-body"> <form method="POST"> {{ form.hidden_tag() }} <div class="mb-3">{{ form.product_id.label(class="form-label") }}{{ form.product_id(class="form-select") }}</div> <div class="mb-3">{{ form.transaction_type.label(class="form-label") }}{{ form.transaction_type(class="form-select") }}</div> <div class="mb-3">{{ form.quantity.label(class="form-label") }}{{ form.quantity(class="form-control") }}</div> <div class="mb-3">{{ form.notes.label(class="form-label") }}{{ form.notes(class="form-control") }}</div> <button type="submit" class="btn btn-success">Registrar</button> <a href="{{ url_for('index') }}" class="btn btn-secondary">Cancelar</a> </form> </div> </div> {% endblock %} """)
write_file("templates/all_transactions.html", """{% extends 'base.html' %}
{% block content %}

<h2>Todos los Movimientos</h2> <a href="{{ url_for('index') }}" class="btn btn-secondary mb-3">← Volver</a> <table class="table table-bordered table-hover"> <thead class="table-dark"> <tr> <th>No.</th> <th>Producto</th> <th>Tipo</th> <th>Cantidad</th> <th>Fecha/Hora</th> <th>Notas</th> <th>Acciones</th> </tr> </thead> <tbody> {% for t in pagination.items %} <tr> <td>{{ loop.index + (pagination.page-1)*pagination.per_page }}</td> <td>{{ t.product.name }}</td> <td><span class="badge bg-{{ 'success' if t.transaction_type=='IN' else 'danger' }}">{{ 'Entrada' if t.transaction_type=='IN' else 'Salida' }}</span></td> <td>{{ t.quantity }}</td> <td>{{ t.transaction_date.strftime('%Y-%m-%d %H:%M:%S') }}</td> <td>{{ t.notes or '-' }}</td> <td> <form method="POST" action="{{ url_for('delete_transaction', id=t.id) }}" style="display:inline;" onsubmit="return confirm('¿Mover a papelera?')"> <button type="submit" class="btn btn-sm btn-danger">🗑️</button> </form> </td> </tr> {% endfor %} </tbody> </table> <nav> <ul class="pagination"> {% if pagination.has_prev %}<li class="page-item"><a class="page-link" href="?page={{ pagination.prev_num }}">Anterior</a></li>{% endif %} <li class="page-item active"><span class="page-link">Página {{ pagination.page }} de {{ pagination.pages }}</span></li> {% if pagination.has_next %}<li class="page-item"><a class="page-link" href="?page={{ pagination.next_num }}">Siguiente</a></li>{% endif %} </ul> </nav> {% endblock %} """)
write_file("templates/deleted_products.html", """{% extends 'base.html' %}
{% block content %}

<h2>🗑️ Papelera de Productos</h2> <a href="{{ url_for('index') }}" class="btn btn-primary mb-3">← Volver</a> <table class="table table-striped"> <thead class="table-dark"> <tr><th>Nombre</th><th>Acciones</th></tr> </thead> <tbody> {% for p in products %} <tr> <td>{{ p.name }}</td> <td><a href="{{ url_for('restore_product', id=p.id) }}" class="btn btn-sm btn-success">↩️ Restaurar</a></td> </tr> {% else %} <tr><td colspan="2" class="text-center">No hay productos en papelera.{% endfor %} </tbody> </table> {% endblock %} """)
write_file("templates/deleted_transactions.html", """{% extends 'base.html' %}
{% block content %}

<h2>🗑️ Movimientos Eliminados</h2> <a href="{{ url_for('index') }}" class="btn btn-primary mb-3">← Volver</a> <table class="table table-striped"> <thead class="table-dark"> <tr><th>Producto</th><th>Tipo</th><th>Cantidad</th><th>Fecha</th><th>Acciones</th></tr> </thead> <tbody> {% for t in pagination.items %} <tr> <td>{{ t.product.name }}</td> <td><span class="badge bg-{{ 'success' if t.transaction_type=='IN' else 'danger' }}">{{ 'Entrada' if t.transaction_type=='IN' else 'Salida' }}</span></td> <td>{{ t.quantity }}</td> <td>{{ t.transaction_date.strftime('%Y-%m-%d %H:%M') }}</td> <td> <form method="POST" action="{{ url_for('restore_transaction', id=t.id) }}" style="display:inline;"> <button type="submit" class="btn btn-sm btn-success">↩️ Restaurar</button> </form> </td> </tr> {% else %} <tr><td colspan="5" class="text-center">No hay movimientos en papelera.{% endfor %} </tbody> </td> <nav> <ul class="pagination"> {% if pagination.has_prev %}<li class="page-item"><a class="page-link" href="?page={{ pagination.prev_num }}">Anterior</a></li>{% endif %} <li class="page-item active"><span class="page-link">Página {{ pagination.page }} de {{ pagination.pages }}</span></li> {% if pagination.has_next %}<li class="page-item"><a class="page-link" href="?page={{ pagination.next_num }}">Siguiente</a></li>{% endif %} </ul> </nav> {% endblock %} """)
write_file("static/css/style.css", """body { background: #f8f9fa; }
.card { border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); }
""")

# ==================== FUNCIONES DE EJECUCIÓN ====================
def run_command(cmd, cwd=None):
    print(f"\n> {cmd}")
    process = subprocess.Popen(cmd, shell=True, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                               text=True, encoding='utf-8', errors='replace')
    for line in process.stdout:
        print(line, end='')
    process.wait()
    if process.returncode != 0:
        print(f"⚠️  El comando falló con código {process.returncode}")
        sys.exit(process.returncode)

def run_application():
    os.chdir(PROJECT_ROOT)
    run_command("python -m venv venv")
    if sys.platform == "win32":
        python_exe = "venv\Scripts\python.exe"
        flask_exe = "venv\Scripts\flask.exe"
    else:
        python_exe = "venv/bin/python"
        flask_exe = "venv/bin/flask"
    run_command(f"{python_exe} -m pip install --upgrade pip")
    run_command(f"{python_exe} -m pip install -r requirements.txt")
    run_command(f"{flask_exe} init-db")
    run_command(f"{flask_exe} import-products")
    print("\n🚀 Levantando servidor Flask...")
    run_command(f"{python_exe} app.py")

if __name__ == '__main__':
    download_static_dependencies()
    print("🚀 Generando proyecto Tarjeta de Estiba (sin valores monetarios)...")
    print("📁 Carpeta: tarjeta_estiba_sin_valores/")
    print(f"✅ {len(PRODUCT_NAMES)} productos importados")
    print("✅ Movimientos sin precios (solo cantidades)")
    print("✅ Exportación a PDF y Excel por producto")
    print("\n🚀 Iniciando instalación automática y ejecución...\n")
    run_application()